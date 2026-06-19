#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""单只股票定增评分 — 终端表格输出(与批量脚本 Excel 结论一致)

复用 batch_screen_and_score 全流程(单行临时 Excel → 筛选 + 财务评分 + 最终结论 + ML),
读取结果 Excel, 只把「单项结论 + 最终结论」以表格打印到终端。跑的是与批量完全相同的
代码路径, 故结论与 Excel 逐字一致(模型版本/特征/阈值全部复用)。

用法:
  python scripts/score_one_stock.py <股票代码> [股票简称] [报价日YYYYMMDD]

  python scripts/score_one_stock.py 300604.SZ 长川科技 20240601   # 完整指定
  python scripts/score_one_stock.py 300604.SZ                     # 名称自动查, 报价日=今天
  python scripts/score_one_stock.py 300604.SZ '' 20240601 --keep  # 保留临时 Excel

选项:
  --keep        保留生成的 scored Excel(复制到当前目录), 便于查看全量明细
"""

import os
import re
import sys
import glob
import shutil
import subprocess
import tempfile
import unicodedata
from datetime import datetime

import pandas as pd

# ── 路径 ──
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
PKG_DIR = os.path.dirname(SCRIPTS_DIR)            # price_maintenance_risk_analysis/
BATCH_SCRIPT = os.path.join(SCRIPTS_DIR, 'batch_screen_and_score.py')

# ── Python 解释器(优先 vnpy, ML/财务评分依赖 py3.10) ──
VNPY_PYTHON = os.path.expanduser('~/anaconda3/envs/vnpy/bin/python')
PYTHON = VNPY_PYTHON if os.path.exists(VNPY_PYTHON) else sys.executable

# 子场景明细列(与 batch_screener._analyze_one 一致)
SUB_SCENARIOS = ['市场指数', '行业PE', '个股PE', 'DCF估值', '修正PE估值', '参数构造', '蒙特卡洛', '反向推算']


def lookup_stock_name(code):
    """通过 tushare 反查股票简称, 失败返回 ''。"""
    try:
        if PKG_DIR not in sys.path:
            sys.path.insert(0, PKG_DIR)
        from tushare_token import resolve_tushare_token
        os.environ.setdefault('TUSHARE_TOKEN', resolve_tushare_token())
        import tushare as ts
        pro = ts.pro_api()
        df = pro.stock_basic(ts_code=code, fields='ts_code,name')
        if df is not None and not df.empty:
            return str(df.iloc[0]['name']).strip()
    except Exception:
        pass
    return ''


def _fmt(v):
    """单元格值 → 展示字符串; None/NaN/'-'占位统一为 '-'。"""
    if v is None:
        return '-'
    s = str(v).strip()
    if not s or s.lower() in ('nan', 'none', 'nat', 'n/a'):
        return '-'
    return s


# ── 终端表格(CJK 等宽对齐) ──
def _cw(s):
    """字符串显示宽度(CJK/全角=2, 其余=1)。"""
    return sum(2 if unicodedata.east_asian_width(ch) in ('W', 'F') else 1 for ch in s)


def _print_table(sections):
    """打印分组对齐表格。sections = [(标题, [(项, 值), ...]), ...], 标题为 None 表示最终结论分隔。"""
    all_items = [it for _, items in sections for it in items]
    key_w = max(_cw(k) for k, _ in all_items) if all_items else 0

    for title, items in sections:
        if title is None:
            print('  ' + '─' * (key_w + 4))
        else:
            print(f"\n【{title}】")
        for k, v in items:
            print(f"  {k}{' ' * (key_w - _cw(k))}   {v}")


def _read_scored_row(scored_path):
    """读取 scored Excel 第 2 行(唯一数据行)→ {列名: 值}。"""
    import openpyxl
    wb = openpyxl.load_workbook(scored_path)
    ws = wb.active
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    vals = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    return dict(zip(header, vals))


def build_sections(row, code, name, issue_date):
    """从数据行构造展示分组。"""
    sections = []

    # ── 定增决策(含子场景明细) ──
    decision_items = [('定增决策', _fmt(row.get('定增决策')))]
    eff = _fmt(row.get('有效阈值数'))
    decision_items.append(('有效阈值数', eff))
    pmin, pmax = _fmt(row.get('溢价率下限')), _fmt(row.get('溢价率上限'))
    decision_items.append(('溢价率区间', f'{pmin} ~ {pmax}' if (pmin != '-' or pmax != '-') else '-'))
    for sub in SUB_SCENARIOS:
        decision_items.append((sub, _fmt(row.get(sub))))
    sections.append(('定增决策', decision_items))

    # ── 财务评分趋势 ──
    fin_items = [
        ('总分趋势', _fmt(row.get('总分_趋势'))),
        ('盈利能力趋势', _fmt(row.get('盈利能力_趋势'))),
        ('成长能力趋势', _fmt(row.get('成长能力_趋势'))),
        ('综合趋势', _fmt(row.get('综合趋势'))),
    ]
    sections.append(('财务评分趋势', fin_items))

    # ── ML 盈利概率(动态发现所有 盈利概率_* / 档位_* 列, 配对输出) ──
    # 7m 评分卡档位(主+BLUE, tag 以"评分卡"开头)参与最终结论门槛: >5 标 ✓, ≤5 标 ✗;
    # 1m/3m 短期列仅展示不标门槛。
    prob_cols = [c for c in row.keys() if str(c).startswith('盈利概率_')]
    ml_items = []
    for pc in prob_cols:
        tag = str(pc)[len('盈利概率_'):]
        tier_raw = _fmt(row.get(f'档位_{tag}'))
        mark = ''
        if tag.startswith('评分卡'):   # 主评分卡 / BLUE(7m), 排除 1m/3m/LGB
            try:
                mark = ' ✓' if int(float(tier_raw)) > 5 else ' ✗'
            except (TypeError, ValueError):
                mark = ' ✗'
        ml_items.append((tag, f'{_fmt(row.get(pc))}   (档位 {tier_raw}{mark})'))
    if ml_items:
        sections.append(('ML 盈利概率', ml_items))

    # ── 最终结论 ──
    final = _fmt(row.get('最终结论'))
    # 精确匹配: "不通过" 含子串"通过", 不能用 in 判断
    mark = '✅ ' if final == '通过' else ('❌ ' if final == '不通过' else '')
    sections.append((None, [('最终结论', mark + final)]))

    return sections


def run_pipeline(code, name, issue_date, keep):
    """跑批量全流程(单行临时输入), 返回 scored Excel 路径。"""
    with tempfile.TemporaryDirectory(prefix='score_one_stock_') as td:
        # 1) 单行输入 Excel
        inp = os.path.join(td, 'input.xlsx')
        pd.DataFrame([{'股票代码': code, '股票简称': name, '报价日': issue_date}]) \
            .to_excel(inp, index=False)

        # 2) 跑 batch_screen_and_score(--force 全量, 单行)
        cmd = [PYTHON, BATCH_SCRIPT, '--input', inp, '--sheet', '0', '--force']
        print(f'\n⏳ 评分中（单只股票，约 30~90 秒，含行业阈值/财报取数）...\n')
        proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, cwd=PKG_DIR, bufsize=1,
        )
        # 过滤冗长的 logger INFO/DEBUG(财务取数日志), 只回显关键进度
        noise = re.compile(r'^\d{4}-\d{2}-\d{2}.*\[(INFO|DEBUG)\]')
        for line in proc.stdout:
            s = line.rstrip()
            if not s or noise.match(s):
                continue
            print('  ' + s)
        proc.wait()
        if proc.returncode != 0:
            print(f'\n❌ 评分流程失败(退出码 {proc.returncode})')
            sys.exit(1)

        # 3) 定位 scored Excel
        hits = glob.glob(os.path.join(td, '*_scored.xlsx'))
        if not hits:
            print('\n❌ 未找到评分结果 Excel(流程可能未产出数据)')
            sys.exit(1)
        scored_path = hits[0]

        # 4) 读取并打印
        row = _read_scored_row(scored_path)
        sections = build_sections(row, code, name, issue_date)

        print('\n' + '═' * 56)
        print(f'  {code}  {name or "(未知)"}    报价日 {issue_date}')
        print('═' * 56)
        _print_table(sections)
        print()

        # 5) 可选: 保留 scored Excel
        if keep:
            dst = os.path.join(os.getcwd(), f'scored_{code.replace(".", "_")}.xlsx')
            shutil.copy(scored_path, dst)
            print(f'  📄 全量明细已保留: {dst}')

    return None


def main():
    import argparse
    ap = argparse.ArgumentParser(
        description='单只股票定增评分 — 终端表格输出(结论与批量 Excel 一致)',
        usage='%(prog)s <股票代码> [股票简称] [报价日YYYYMMDD] [--keep]',
    )
    ap.add_argument('code', help='股票代码(如 300604.SZ)')
    ap.add_argument('name', nargs='?', default=None, help='股票简称(可选, 缺省自动查询)')
    ap.add_argument('issue_date', nargs='?', default=None, help='报价日 YYYYMMDD(可选, 缺省=今天)')
    ap.add_argument('--keep', action='store_true', help='保留生成的 scored Excel 到当前目录(查看全量明细)')
    args = ap.parse_args()

    code = args.code.strip()
    issue_date = (args.issue_date or datetime.now().strftime('%Y%m%d')).strip()

    # 名称: 未给则查
    name = (args.name or '').strip()
    if not name:
        name = lookup_stock_name(code)
        if name:
            print(f'  名称自动查询: {name}')

    run_pipeline(code, name, issue_date, args.keep)


if __name__ == '__main__':
    main()
