#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量财务评分脚本
读取上市公司清单，自动计算近5年财务评分，结果写回Excel

用法:
    python price_maintenance_risk_analysis/scripts/batch_financial_score.py
    python scripts/batch_financial_score.py <excel_path>
"""

import sys
import os
import time
import threading
import logging

import pandas as pd
import numpy as np
import tushare as ts
import openpyxl
import warnings

warnings.filterwarnings('ignore', category=FutureWarning)

_EFAES_ROOT = os.path.expanduser('~/github/EFAES')
sys.path.insert(0, _EFAES_ROOT)

from src.web.config import DB_CONFIG, TUSHARE_TOKEN
from src.core.tushare_data_fetcher import TushareDataFetcher
from src.core.industry_threshold_calculator import IndustryThresholdCalculator
from src.core.calculate_financial_score import FinancialScoringCard
from src.core.calculate_financial_ratios import calculate_financial_ratios

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

CURRENT_YEAR = 2026
SCORE_YEARS = list(range(CURRENT_YEAR - 5, CURRENT_YEAR))  # 默认，会被每只股票的报价日覆盖
RELATIVE_YEAR_LABELS = ['T-4', 'T-3', 'T-2', 'T-1', 'T']  # 相对年份表头

# 行业阈值缓存，同行业不重复计算
_industry_thresholds_cache = {}  # {(industry, year): thresholds}
_industry_raw_cache = {}  # {industry: (df_with_stats, fetch_end_year)} 原始数据缓存

# 行业查询节流锁：多线程并发查 index_member_all 会打爆 tushare 限频，
# 限频后接口返回空 → 行业缺失 → 整行评分被跳过。用此锁把行业查询串行化 + 每次间隔，
# 配合下面的重试，避免限频导致的评分缺失。
_industry_query_lock = threading.Lock()

# 二/一级行业回退阈值缓存：冷门三级行业样本不足时回退到二级/一级，同行业只算一次
_fallback_thresholds_cache = {}  # {(industry_name, year): thresholds}


def get_company_industry(pro, ts_code: str, max_retries: int = 3) -> dict:
    """通过 Tushare 反查申万三级行业，返回 {l1_name, l2_name, l3_name}

    带节流+重试：index_member_all 接口在并发查询时易触发限频返回空，
    用全局锁串行化调用 + 每次间隔，并在返回空/异常时退避重试，避免限频导致评分缺失。
    """
    result = {'l1_name': '', 'l2_name': '', 'l3_name': ''}
    for attempt in range(max_retries):
        try:
            # 节流：串行化 index_member_all 调用 + 间隔，避免打爆限频
            with _industry_query_lock:
                df = pro.index_member_all(ts_code=ts_code)
                time.sleep(0.35)
            if df is not None and not df.empty:
                row = df.sort_values('in_date', ascending=False).iloc[0]
                l1 = row.get('l1_name', '')
                l2 = row.get('l2_name', '')
                l3 = row.get('l3_name', '')
                if l3:
                    logger.info(f"{ts_code} 申万行业: {l1} > {l2} > {l3}")
                    return {'l1_name': l1, 'l2_name': l2, 'l3_name': l3}
                # df 非空但 l3 空：数据本身无三级分类，不重试
                return {'l1_name': l1, 'l2_name': l2, 'l3_name': l3}
            # df 为空：疑似限频，退避后重试
            logger.warning(f"{ts_code} 行业查询返回空(尝试 {attempt + 1}/{max_retries})，疑似限频")
        except Exception as e:
            logger.warning(f"查询 {ts_code} 申万行业失败(尝试 {attempt + 1}/{max_retries}): {e}")
        if attempt < max_retries - 1:
            time.sleep(1.0 * (attempt + 1))  # 退避：1s, 2s

    logger.warning(f"{ts_code} 行业查询 {max_retries} 次均失败/为空，评分将被跳过")
    return result


def _get_industry_stocks_by_level(pro, industry_name, level='L2'):
    """获取申万 L2/L1 行业的成分股列表（阈值兜底用）。

    EFAES 的 get_industry_stocks 只认三级行业名（硬编码 level='L3'）；
    本函数补充 L2/L1，用于三级行业样本不足时回退。返回 ts_code 列表，失败返回 []。
    """
    try:
        df = pro.index_classify(level=level, src='SW2021')
        if df is None or df.empty:
            return []
        matched = df[df['industry_name'] == industry_name]
        if matched.empty:
            matched = df[df['industry_name'].str.contains(industry_name, na=False)]
        if matched.empty:
            logger.warning(f"申万{level} 行业 '{industry_name}' 未匹配到指数")
            return []
        index_code = matched.iloc[0]['index_code']
        logger.info(f"申万{level}匹配: '{industry_name}' -> {index_code}")
        members = pro.index_member(index_code=index_code, is_new='Y')
        if members is None or members.empty:
            return []
        return members['con_code'].unique().tolist()
    except Exception as e:
        logger.warning(f"申万{level}行业 '{industry_name}' 成分股查询失败: {e}")
        return []


def fetch_financial_data(fetcher: TushareDataFetcher, ts_code: str, score_years=None, max_retries: int = 3):
    """获取公司三张报表数据，范围覆盖score_years（需前一年做同比）。

    带重试：fina_indicator/财报接口在并发下易触发 tushare 限频(500次/分钟)返回空，
    空时退避重试，避免限频导致个股三表为空、评分被跳过。
    """
    years = score_years if score_years else SCORE_YEARS
    start_date = f'{min(years) - 1}0101'
    end_date = f'{max(years)}1231'

    def _try(fetch_fn, label):
        for attempt in range(max_retries):
            try:
                df = fetch_fn(ts_code, start_date, end_date)
                if df is not None and not df.empty:
                    return df
            except Exception as e:
                logger.warning(f"{ts_code} {label}获取异常(尝试 {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(1.0 * (attempt + 1))  # 退避：1s, 2s
        return pd.DataFrame()

    bs = _try(fetcher.fetch_balance_sheet, '资产负债表')
    inc = _try(fetcher.fetch_income_statement, '利润表')
    cf = _try(fetcher.fetch_cash_flow_statement, '现金流量表')

    for df in [bs, inc, cf]:
        if not df.empty:
            df['report_year'] = df['end_date'].apply(lambda x: int(str(x)[:4]) if pd.notna(x) else None)
            company_name = fetcher._get_company_name(ts_code) or ts_code
            df['company_name'] = company_name

    return bs, inc, cf


def get_industry_thresholds(threshold_calc: IndustryThresholdCalculator, industry_name: str,
                            target_year: int = None) -> dict:
    """获取行业阈值（按年份缓存），优先用target_year，不足则往前找。

    Args:
        target_year: 目标年份（报价日年份），阈值反映该年行业分布。
                     None则用SCORE_YEARS[-1]。
    """
    global _industry_thresholds_cache

    # 按行业+年份组合缓存（不同年份行业阈值不同）
    cache_key = (industry_name, target_year)
    if cache_key in _industry_thresholds_cache:
        logger.info(f"使用缓存的 '{industry_name}' {target_year}年 行业阈值")
        return _industry_thresholds_cache[cache_key]

    # 确定尝试的年份：从target_year往前找（最多回退3年）
    base_year = target_year if target_year else SCORE_YEARS[-1]
    try_years = [base_year - i for i in range(4)]  # target_year, T-1, T-2, T-3

    # 原始数据缓存：同一行业只取一次数，范围覆盖所有可能的报价日年份
    # start_date要足够早（最早报价日2020的前5年=2015），end_date用CURRENT_YEAR
    fetch_start_year = min(base_year - 5, 2010)  # 至少回溯到报价日前5年
    fetch_end_year = CURRENT_YEAR
    df_with_stats = None
    if industry_name in _industry_raw_cache:
        cached_stats, cached_start, cached_end = _industry_raw_cache[industry_name]
        # 缓存数据需同时覆盖start和end
        if cached_start <= fetch_start_year and cached_end >= base_year:
            df_with_stats = cached_stats
            logger.info(f"复用 '{industry_name}' 已缓存的原始数据（{cached_start}~{cached_end}年）")

    if df_with_stats is None:
        start_date = f'{fetch_start_year}0101'
        end_date = f'{fetch_end_year}1231'
        try:
            logger.info(f"获取 '{industry_name}' 行业原始数据（{start_date}~{end_date}）...")
            ts_codes = threshold_calc.get_industry_stocks(industry_name)
            if not ts_codes:
                logger.error(f"行业 '{industry_name}' 无成分股")
                _industry_thresholds_cache[cache_key] = {}
                return {}

            # 同时传入start_date和end_date，确保覆盖所有年份
            df_fina = threshold_calc.fetch_financial_data(ts_codes, start_date=start_date, end_date=end_date)
            if df_fina.empty:
                logger.error(f"行业 '{industry_name}' 财务数据为空")
                _industry_thresholds_cache[cache_key] = {}
                return {}

            df_income = threshold_calc.fetch_rd_expense_data(ts_codes, end_date=end_date)
            df_processed = threshold_calc.process_financial_data(df_fina, df_income)
            from src.core.tushare_tools import add_industry_statistics
            df_with_stats = add_industry_statistics(df_processed)
            _industry_raw_cache[industry_name] = (df_with_stats, fetch_start_year, fetch_end_year)
            logger.info(f"✅ '{industry_name}' 原始数据已缓存（{len(df_with_stats)}条, {fetch_start_year}~{fetch_end_year}），后续年份复用")
        except Exception as e:
            logger.error(f"获取 '{industry_name}' 原始数据失败: {e}")
            _industry_thresholds_cache[cache_key] = {}
            return {}

    # 从已缓存的原始数据计算各年阈值
    thresholds = {}
    for try_year in try_years:
        try:
            result = threshold_calc.create_thresholds_from_dataframe(df_with_stats, try_year, industry_name)
            if not result or 'thresholds' not in result:
                continue
            t = result['thresholds']
            valid = sum(1 for v in t.values() if v is not None)
            if valid >= 10:
                logger.info(f"行业 '{industry_name}' 使用 {try_year} 年阈值（有效指标: {valid}）")
                thresholds = t
                break
        except Exception as e:
            logger.warning(f"从缓存数据计算 {try_year} 年阈值失败: {e}")
            continue

    if not thresholds:
        logger.error(f"计算行业 '{industry_name}' 的阈值失败")
        return {}

    clean_thresholds = {k: v for k, v in thresholds.items() if v is not None}
    _industry_thresholds_cache[cache_key] = clean_thresholds
    return clean_thresholds


def _calc_thresholds_for_codes(threshold_calc, ts_codes, industry_name, target_year=None):
    """用外部成分股 ts_codes 计算行业阈值（二/一级 fallback 用），带年份回退与内存缓存。

    复用 EFAES calculate_custom_industry_thresholds（接收外部 ts_codes，绕开只认三级的
    get_industry_stocks）。冷门三级行业样本不足时，用其所属二/一级的成分股算阈值兜底。
    """
    base_year = target_year if target_year else SCORE_YEARS[-1]
    for try_year in [base_year - i for i in range(4)]:
        cache_key = (industry_name, try_year)
        if cache_key in _fallback_thresholds_cache:
            logger.info(f"复用回退行业 '{industry_name}' {try_year}年阈值(缓存)")
            return _fallback_thresholds_cache[cache_key]
        try:
            result = threshold_calc.calculate_custom_industry_thresholds(
                industry_name, ts_codes, target_year=try_year, save_to_db=False)
            if not result or 'thresholds' not in result:
                continue
            t = result.get('thresholds') or {}
            valid = sum(1 for v in t.values() if v is not None)
            if valid >= 10:
                clean = {k: v for k, v in t.items() if v is not None}
                logger.info(f"回退行业 '{industry_name}' 使用 {try_year} 年阈值（有效指标 {valid}/{len(t)}）")
                _fallback_thresholds_cache[cache_key] = clean
                return clean
        except Exception as e:
            logger.warning(f"回退行业 '{industry_name}' {try_year}年阈值计算失败: {e}")
            continue
    return {}


def _save_financial_indicators(stock_code, report_year, ratios):
    """将计算好的财务指标保存到 investment_valuation.financial_indicators 表"""
    import pymysql

    # 指标字段列表(与表结构一一对应)
    indicator_fields = [
        'current_ratio', 'quick_ratio', 'inv_turn', 'ar_turn', 'ca_turn', 'assets_turn',
        'roa', 'npta', 'roe', 'roe_dt', 'netprofit_margin', 'grossprofit_margin',
        'debt_to_assets', 'int_to_talcap', 'debt_to_eqt', 'ebit_to_interest',
        'cash_to_liqdebt', 'cash_to_liqdebt_withinterest', 'rd_exp_ratio',
        'op_yoy', 'ebt_yoy', 'netprofit_yoy', 'dt_netprofit_yoy', 'roe_yoy',
        'tr_yoy', 'or_yoy', 'equity_yoy',
    ]

    # 构建 VALUES 部分
    values = [stock_code, report_year]
    placeholders = ['%s', '%s']
    update_parts = []
    for f in indicator_fields:
        values.append(ratios.get(f))
        placeholders.append('%s')
        update_parts.append(f'{f}=VALUES({f})')

    values.append(1 if ratios.get('used_average_values') else 0)
    placeholders.append('%s')
    update_parts.append('used_average_values=VALUES(used_average_values)')

    values.append(ratios.get('valid_indicator_count', 0))
    placeholders.append('%s')
    update_parts.append('valid_indicator_count=VALUES(valid_indicator_count)')

    conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='',
                           database='investment_valuation', charset='utf8mb4')
    try:
        with conn.cursor() as cur:
            cols = ', '.join(['stock_code', 'report_year'] + indicator_fields +
                             ['used_average_values', 'valid_indicator_count'])
            ph = ', '.join(placeholders)
            sql = f"INSERT INTO financial_indicators ({cols}) VALUES ({ph}) ON DUPLICATE KEY UPDATE {', '.join(update_parts)}"
            cur.execute(sql, values)
            conn.commit()
    finally:
        conn.close()


def _save_annual_scores_to_db(stock_code, results, stock_years):
    """将年度财务评分保存到 investment_valuation.company_annual_scores"""
    import sys as _sys
    _proj_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _proj_root not in _sys.path:
        _sys.path.insert(0, _proj_root)
    from utils.db_manager import ValuationDB
    db = ValuationDB()

    scores_by_year = {}
    for year in stock_years:
        if year in results and isinstance(results[year], dict):
            r = results[year]
            scores_by_year[year] = {
                'stock_name': None,
                'total_score': r.get('总分'),
                'rating': r.get('评级'),
                'profitability': r.get('盈利能力'),
                'growth': r.get('成长能力'),
                'operating': r.get('运营能力'),
                'solvency': r.get('偿债能力'),
            }

    # 补充行业信息
    industry_info = results.get('行业') or {}
    for year_data in scores_by_year.values():
        year_data['industry_l1'] = industry_info.get('l1_name')
        year_data['industry_l2'] = industry_info.get('l2_name')
        year_data['industry_l3'] = industry_info.get('l3_name')

    if scores_by_year:
        db.save_annual_scores(stock_code, scores_by_year)


def score_company(fetcher: TushareDataFetcher, threshold_calc: IndustryThresholdCalculator,
                  ts_code: str, company_name: str, score_years=None) -> dict:
    """计算单个公司近5年的财务评分

    Args:
        score_years: 评分年份列表（从报价日回溯），None则用全局SCORE_YEARS
    """
    pro = fetcher.pro
    years_to_score = score_years if score_years else SCORE_YEARS
    logger.info(f"{'='*60}")
    logger.info(f"开始处理: {ts_code} {company_name}")
    logger.info(f"{'='*60}")

    # 1. 获取行业
    industry_info = get_company_industry(pro, ts_code)
    if not industry_info.get('l3_name'):
        logger.error(f"无法确定 {ts_code} 的三级行业，跳过")
        return {}

    # 2. 获取行业阈值：三级 → 二级 → 一级 回退（冷门三级行业样本不足时兜底）
    industry_name = industry_info['l3_name']
    thresholds = get_industry_thresholds(threshold_calc, industry_name)
    if not thresholds:
        for level_name, key, sw_level in (('二级', 'l2_name', 'L2'), ('一级', 'l1_name', 'L1')):
            fb_name = industry_info.get(key, '')
            if not fb_name:
                continue
            fb_codes = _get_industry_stocks_by_level(pro, fb_name, level=sw_level)
            if not fb_codes:
                continue
            thresholds = _calc_thresholds_for_codes(threshold_calc, fb_codes, fb_name)
            if thresholds:
                logger.info(f"{ts_code} 三级行业 '{industry_name}' 阈值样本不足，回退{level_name}行业 '{fb_name}'({len(fb_codes)}只成分股)")
                industry_name = fb_name
                break
    if not thresholds:
        logger.error(f"{ts_code} 三/二/一级行业阈值均计算失败，跳过")
        return {}

    logger.info(f"行业阈值加载完成（{industry_name}），共 {len(thresholds)} 个指标")
    valid_count = sum(1 for v in thresholds.values() if v is not None)
    logger.info(f"有效阈值: {valid_count}/{len(thresholds)}")

    # 3. 获取财务数据（按各自报价日年份范围查询）
    bs, inc, cf = fetch_financial_data(fetcher, ts_code, score_years=years_to_score)
    if bs.empty or inc.empty:
        logger.error(f"{ts_code} 财务数据为空，跳过")
        return {}

    # 4. 逐年计算评分
    results = {}
    scorer = FinancialScoringCard(DB_CONFIG)
    scorer.set_industry_thresholds(thresholds)

    for year in years_to_score:
        try:
            bs_year = bs[bs['report_year'] == year]
            inc_year = inc[inc['report_year'] == year]

            if bs_year.empty or inc_year.empty:
                logger.warning(f"{ts_code} {year}年 报表数据缺失，跳过")
                continue

            ratios = calculate_financial_ratios(bs, inc, cf, year)
            if not ratios:
                logger.warning(f"{ts_code} {year}年 指标计算失败，跳过")
                continue

            # 落库: 保存财务指标到 investment_valuation.financial_indicators
            try:
                _save_financial_indicators(ts_code, year, ratios)
            except Exception as e:
                logger.debug(f"{ts_code} {year}年 指标落库失败(不影响评分): {e}")

            scorer.load_company_data(ratios, company_name)
            score_result = scorer.calculate_score()

            total_score = score_result['总得分']
            rating = scorer.get_rating(total_score)
            dim_scores = score_result['维度得分']

            results[year] = {
                '总分': round(total_score, 2),
                '评级': rating,
                '运营能力': round(dim_scores.get('运营能力', 0), 2),
                '盈利能力': round(dim_scores.get('盈利能力', 0), 2),
                '偿债能力': round(dim_scores.get('偿债能力', 0), 2),
                '成长能力': round(dim_scores.get('成长能力', 0), 2),
            }
            logger.info(f"  {year}年: 总分={total_score:.2f} 评级={rating}")

        except Exception as e:
            logger.warning(f"{ts_code} {year}年 评分计算失败: {e}")
            import traceback
            traceback.print_exc()
            continue

    results['行业'] = industry_info
    return results


def main():
    import argparse
    parser = argparse.ArgumentParser(description='批量财务评分（支持行级增量续跑）')
    parser.add_argument('excel_path', help='输入Excel（screening结果）；若对应 _scored 已存在则续传补算')
    parser.add_argument('--force', action='store_true', help='强制全量重算（忽略已有 _scored，从头评分）')
    args = parser.parse_args()

    excel_path = args.excel_path
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    # 续传模式：若 _scored 已存在且非 --force，则读它（保留已评分行，只补失败行）
    output_path = excel_path.replace('.xlsx', '_scored.xlsx')
    resume_mode = (not args.force) and os.path.exists(output_path)
    src_path = output_path if resume_mode else excel_path
    logger.info(f"读取 Excel: {src_path}" + ("（行级续传：仅补算评分失败的行）" if resume_mode else "（全量评分）"))

    # 初始化
    pro = ts.pro_api(TUSHARE_TOKEN)
    fetcher = TushareDataFetcher(TUSHARE_TOKEN, DB_CONFIG)
    threshold_calc = IndustryThresholdCalculator(TUSHARE_TOKEN, DB_CONFIG)

    # 读取 Excel
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active

    # 找到股票代码列
    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    code_col = header_row.index('股票代码') + 1
    name_col = header_row.index('股票简称') + 1

    # 读取报价日列（如有），按报价日年份回溯确定评分年份
    quote_date_col = None
    if '报价日' in header_row:
        quote_date_col = header_row.index('报价日') + 1
    # 取第一只有报价日的股票确定评分年份范围
    base_year = CURRENT_YEAR
    if quote_date_col:
        for row_idx in range(2, ws.max_row + 1):
            qd = ws.cell(row_idx, quote_date_col).value
            if qd and str(qd).strip() and len(str(qd).strip()) >= 4:
                try:
                    base_year = int(str(qd).strip()[:4])
                    logger.info(f"检测到报价日 {qd}，评分年份按报价日({base_year}年)回溯")
                    break
                except ValueError:
                    pass
    score_years = list(range(base_year - 4, base_year + 1))  # 报价年及前4年
    logger.info(f"评分年份: {score_years}")
    # 更新全局SCORE_YEARS，使所有函数(get_industry_thresholds/score_company)同步使用
    global SCORE_YEARS
    SCORE_YEARS = score_years

    # 表头定位/初始化（幂等：已有评分列则复用列索引，否则新增——避免续跑重复加一整套列）
    fields = ['总分', '评级', '盈利能力', '成长能力']
    check_fields_trend = ['总分', '盈利能力', '成长能力']

    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if '三级行业' in header_row and '总分_T-4' in header_row and '总分_斜率' in header_row:
        # 续传：复用已有评分列（列顺序与首次写入一致）
        start_col = header_row.index('一级行业') + 1
        data_start_col = header_row.index('总分_T-4') + 1
        trend_start = header_row.index('总分_斜率') + 1
        logger.info("检测到已有评分列，复用列位置（续传模式）")
    else:
        # 首次：追加表头
        start_col = ws.max_column + 1
        ws.cell(1, start_col, '一级行业')
        ws.cell(1, start_col + 1, '二级行业')
        ws.cell(1, start_col + 2, '三级行业')
        data_start_col = start_col + 3

        headers = [f'{field}_{label}' for field in fields for label in RELATIVE_YEAR_LABELS]
        for i, h in enumerate(headers):
            ws.cell(1, data_start_col + i, h)

        trend_start = data_start_col + len(headers)
        trend_headers = []
        for field in check_fields_trend:
            trend_headers.append(f'{field}_斜率')
            trend_headers.append(f'{field}_趋势')
        trend_headers.append('综合趋势')
        for i, h in enumerate(trend_headers):
            ws.cell(1, trend_start + i, h)

    # 刷新表头（首次模式下含新加列；续传模式下不变）
    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    # ===== 并发评分（多线程，API调用为主，适合I/O并行）=====
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _write_lock = threading.Lock()
    _done_count = [0]
    # _total 在下方确定待补算行（row_indices）后设置

    def _score_one(row_idx):
        """评分单只股票（线程内），返回(row_idx, results, score_years)"""
        ts_code = ws.cell(row_idx, code_col).value
        company_name = ws.cell(row_idx, name_col).value
        if not ts_code:
            return row_idx, None, None
        ts_code = str(ts_code).strip()
        # 读取该股票的报价日，计算各自回溯的评分年份
        stock_years = SCORE_YEARS  # 默认
        if quote_date_col:
            qd = ws.cell(row_idx, quote_date_col).value
            if qd and str(qd).strip() and len(str(qd).strip()) >= 4:
                try:
                    base_year = int(str(qd).strip()[:4])
                    stock_years = list(range(base_year - 4, base_year + 1))
                except ValueError:
                    pass
        results = score_company(fetcher, threshold_calc, ts_code, company_name, score_years=stock_years)
        return row_idx, results, stock_years

    def _write_results(row_idx, results, stock_years):
        """将评分结果写入Excel（线程安全，按相对年份顺序写入）"""
        if results is None:
            return
        # 写入三级行业信息
        if results and results.get('行业'):
            info = results['行业']
            ws.cell(row_idx, start_col, info.get('l1_name', ''))
            ws.cell(row_idx, start_col + 1, info.get('l2_name', ''))
            ws.cell(row_idx, start_col + 2, info.get('l3_name', ''))
            results.pop('行业')

        # 按stock_years顺序写入（对应表头T-4,T-3,...,T）
        col_offset = data_start_col
        for field in fields:
            for year in stock_years:
                if year in results:
                    ws.cell(row_idx, col_offset, results[year].get(field, 'N/A'))
                else:
                    ws.cell(row_idx, col_offset, 'N/A')
                col_offset += 1

        if results:
            check_fields_t = ['总分', '盈利能力', '成长能力']
            thresholds = {'总分': -2, '盈利能力': -1, '成长能力': -1}
            fail_count = 0
            for field in check_fields_t:
                scores = [results[y][field] for y in stock_years if y in results and isinstance(results[y].get(field), (int, float))]
                slope = np.polyfit(range(len(scores)), scores, 1)[0] if len(scores) >= 2 else 0
                slope = round(slope, 2)
                status = '不通过' if slope < thresholds[field] else '通过'
                if status == '不通过':
                    fail_count += 1
                ws.cell(row_idx, col_offset, slope)
                ws.cell(row_idx, col_offset + 1, status)
                col_offset += 2

            latest_year = max(results.keys()) if results else 0
            latest_score = results.get(latest_year, {}).get('总分', 0)
            if isinstance(latest_score, (int, float)) and latest_score < 55:
                final = '不通过'
            else:
                final = '不通过' if fail_count >= 2 else '通过'
            ws.cell(row_idx, col_offset, final)

            # ── 落库: 保存年度评分到 company_annual_scores ──
            try:
                _save_annual_scores_to_db(ts_code, results, stock_years)
            except Exception as e:
                logger.debug(f'{ts_code} 年度评分落库失败(不影响评分): {e}')

    # 行级续传：判定待补算行（三级行业为空 或 总分_T 缺失）；全量模式则所有行
    all_rows = list(range(2, ws.max_row + 1))
    l3_col = header_row.index('三级行业') + 1
    total_t_col = header_row.index('总分_T') + 1

    if resume_mode:
        def _needs_rescore(r):
            l3 = ws.cell(r, l3_col).value
            tt = ws.cell(r, total_t_col).value
            l3_empty = l3 is None or str(l3).strip() in ('', 'N/A', 'nan', 'None')
            tt_empty = tt is None or str(tt).strip() in ('', 'N/A', 'nan', 'None')
            return l3_empty or tt_empty
        row_indices = [r for r in all_rows if _needs_rescore(r)]
        logger.info(f"行级续传：已评分 {len(all_rows) - len(row_indices)} 行跳过，待补算 {len(row_indices)} 行")
    else:
        row_indices = all_rows
        logger.info(f"全量评分：共 {len(row_indices)} 行")

    if not row_indices:
        logger.info("✅ 无待补算行，直接保存退出（文件无变更）")
        wb.save(output_path)
        print(f"\n完成! 结果文件: {output_path} (无变更)")
        return

    _total = len(row_indices)
    MAX_WORKERS_SCORE = min(5, len(row_indices))
    logger.info(f"并发评分：{MAX_WORKERS_SCORE}线程并行")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS_SCORE) as executor:
        futures = {executor.submit(_score_one, ridx): ridx for ridx in row_indices}
        for future in as_completed(futures):
            ridx = futures[future]
            try:
                row_idx, results, stock_years = future.result()
                with _write_lock:
                    _write_results(row_idx, results, stock_years)
                    _done_count[0] += 1
                    n = _done_count[0]
                    if n % 10 == 0 or n == _total:
                        output_path = excel_path.replace('.xlsx', '_scored.xlsx')
                        try:
                            wb.save(output_path)
                        except Exception:
                            pass
                    ts_code = str(ws.cell(row_idx, code_col).value or '').strip()
                    logger.info(f"  [已评分 {n}/{_total}] {ts_code}")
            except Exception as e:
                logger.warning(f"  第{ridx}行评分失败: {e}")

    # 最终保存
    output_path = excel_path.replace('.xlsx', '_scored.xlsx')
    wb.save(output_path)
    logger.info(f"\n评分结果已保存: {output_path}")
    print(f"\n完成! 结果文件: {output_path}")


if __name__ == '__main__':
    main()
