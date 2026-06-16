#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量定增筛选 + 财务评分 一体化脚本

串联 batch_screener.py（定增决策筛选）和 batch_financial_score.py（财务评分），
输出带评分的最终结果文件。

用法:
  python scripts/batch_screen_and_score.py --input stocks.xlsx [--sheet 0]

流程:
  Step 1: batch_screener.py  → data/batch_screening_result_<date>.xlsx
  Step 2: batch_financial_score.py → data/batch_screening_result_<date>_scored.xlsx
  Step 3: 生成最终结论列
  Step 4: ML模型预测盈利概率（LightGBM + 逻辑回归）

最终输出: data/batch_screening_result_<date>_scored.xlsx
  最后两列: 盈利概率_LightGBM, 盈利概率_逻辑回归
"""

import os
import sys
import time
import subprocess
from datetime import datetime

import openpyxl


# ── 脚本路径 ──
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
SCREENER_SCRIPT = os.path.join(SCRIPTS_DIR, 'batch_screener.py')
EFAES_SCRIPT = os.path.join(SCRIPTS_DIR, 'batch_financial_score.py')
PREDICT_SCRIPT = os.path.join(os.path.dirname(SCRIPTS_DIR), 'ml_training', 'predict_profitability.py')

DATA_DIR = os.path.join(os.path.dirname(SCRIPTS_DIR), 'data')

# ── Python 解释器（优先 vnpy 环境） ──
VNPY_PYTHON = os.path.expanduser('~/anaconda3/envs/vnpy/bin/python')
PYTHON = VNPY_PYTHON if os.path.exists(VNPY_PYTHON) else sys.executable


def run_step(description, cmd):
    """运行一个子步骤，打印分隔线和结果"""
    print()
    print('=' * 60)
    print(f'  {description}')
    print('=' * 60)
    print(f'  命令: {" ".join(cmd)}')
    print()

    t0 = time.time()
    result = subprocess.run(cmd, cwd=DATA_DIR)
    elapsed = time.time() - t0

    if result.returncode != 0:
        print(f'\n❌ {description} 失败 (退出码: {result.returncode}) [耗时 {elapsed:.1f}s]')
        return False

    print(f'\n✅ {description} 完成 [耗时 {elapsed:.1f}s]')
    return True


def main():
    import argparse

    parser = argparse.ArgumentParser(
        description='批量定增筛选 + 财务评分 一体化',
        usage='%(prog)s --input stocks.xlsx [--sheet 0]',
    )
    parser.add_argument('--input', required=True, help='输入Excel文件路径（含"股票代码"和"股票简称"列）')
    parser.add_argument('--sheet', default='0', help='读取Excel的第几个sheet（序号从0开始，默认0）')
    parser.add_argument('--force', action='store_true', help='强制全量重跑(忽略已有结果, 默认断点续传)')
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f'❌ 文件不存在: {args.input}')
        sys.exit(1)

    if not os.path.exists(SCREENER_SCRIPT):
        print(f'❌ 找不到 batch_screener.py: {SCREENER_SCRIPT}')
        sys.exit(1)

    if not os.path.exists(EFAES_SCRIPT):
        print(f'❌ 找不到 batch_financial_score.py: {EFAES_SCRIPT}')
        sys.exit(1)

    # 生成带日期+输入文件名的输出文件名（输出到输入文件所在目录）
    date_tag = datetime.now().strftime('%y%m%d')
    input_basename = os.path.splitext(os.path.basename(args.input))[0]
    input_dir = os.path.dirname(os.path.abspath(args.input))
    screening_output = os.path.join(input_dir, f'batch_screening_result_{date_tag}_{input_basename}.xlsx')
    scored_output = screening_output.replace('.xlsx', '_scored.xlsx')

    print(f'📁 输入文件: {args.input}')
    print(f'📁 筛选结果: {screening_output}')
    print(f'📁 最终结果: {scored_output}')

    t_total = time.time()  # 总计时起点
    skip_mode = not args.force   # --force 强制全量重跑; 默认断点续传(已有结果跳过)

    # ── Step 1: 定增决策筛选 (已有结果则跳过, 断点续传) ──
    if skip_mode and os.path.exists(screening_output):
        print(f'  ✅ Step 1 已完成(文件存在)，跳过')
    else:
        step1_cmd = [PYTHON, SCREENER_SCRIPT, '--input', args.input,
                     '--output', screening_output, '--sheet', args.sheet]
        if not run_step('Step 1/4: 定增决策筛选 (batch_screener)', step1_cmd):
            sys.exit(1)
        if not os.path.exists(screening_output):
            print(f'❌ 筛选结果文件未生成: {screening_output}'); sys.exit(1)

    # ── Step 2: 财务评分 (已有结果则跳过) ──
    if skip_mode and os.path.exists(scored_output):
        print(f'  ✅ Step 2 已完成(文件存在)，跳过')
    else:
        if not run_step('Step 2/4: 财务评分 (batch_financial_score)', [PYTHON, EFAES_SCRIPT, screening_output]):
            sys.exit(1)
        if not os.path.exists(scored_output):
            print(f'⚠️ 评分结果文件未生成: {scored_output}'); sys.exit(1)

    # ── Step 3: 追加最终结论列 (幂等: 已有则复用列, 不重复添加) ──
    # 规则: 定增决策="建议参与" 且 成长能力_趋势="通过" → 最终"通过"
    t_step3 = time.time()
    print()
    print('=' * 60)
    print('  Step 3/4: 生成最终结论')
    print('=' * 60)

    wb = openpyxl.load_workbook(scored_output)
    ws = wb.active

    # 定位关键列
    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    decision_col = header_row.index('定增决策') + 1
    growth_col = header_row.index('成长能力_趋势') + 1

    # 最终结论列: 已有则复用(幂等), 无则新增
    if '最终结论' in header_row:
        final_col = header_row.index('最终结论') + 1
    else:
        final_col = ws.max_column + 1
        ws.cell(1, final_col, '最终结论')

    pass_count = 0
    for row_idx in range(2, ws.max_row + 1):
        decision = str(ws.cell(row_idx, decision_col).value or '')
        growth = str(ws.cell(row_idx, growth_col).value or '')

        # 定增决策: 包含"建议参与"即为通过
        decision_pass = '建议参与' in decision
        # 成长能力趋势: 精确匹配"通过"
        growth_pass = growth.strip() == '通过'

        conclusion = '通过' if (decision_pass and growth_pass) else '不通过'
        ws.cell(row_idx, final_col, conclusion)
        if conclusion == '通过':
            pass_count += 1

        stock_name = ws.cell(row_idx, header_row.index('股票简称') + 1).value or ''
        stock_code = ws.cell(row_idx, header_row.index('股票代码') + 1).value or ''
        print(f'  {stock_code} {stock_name}: 定增={"✓" if decision_pass else "✗"} 成长={"✓" if growth_pass else "✗"} → {conclusion}')

    wb.save(scored_output)

    # ── 落库: 将 scored Excel 数据写入 placement_evaluation ──
    t_db = time.time()
    print()
    print('  落库: 保存评估记录到 placement_evaluation...')
    try:
        db_dir = os.path.join(os.path.dirname(SCRIPTS_DIR), 'utils')
        if db_dir not in sys.path:
            sys.path.insert(0, db_dir)
        from db_manager import ValuationDB
        db = ValuationDB()

        # 重新读header（可能已被scoring追加列）
        header_row_full = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        code_col_idx = header_row_full.index('股票代码') + 1
        name_col_idx = header_row_full.index('股票简称') + 1

        # 辅助: 安全取列index
        def _col_idx(name, default=None):
            return header_row_full.index(name) + 1 if name in header_row_full else default

        quote_date_ci = _col_idx('报价日')
        quote_price_ci = _col_idx('报价日价格')
        thresholds_ci = _col_idx('有效阈值数')
        prem_min_ci = _col_idx('溢价率下限')
        prem_max_ci = _col_idx('溢价率上限')
        decision_ci = _col_idx('定增决策')
        l1_ci = _col_idx('一级行业')
        l2_ci = _col_idx('二级行业')
        l3_ci = _col_idx('三级行业')
        total_slope_ci = _col_idx('总分_斜率')
        total_trend_ci = _col_idx('总分_趋势')
        profit_slope_ci = _col_idx('盈利能力_斜率')
        profit_trend_ci = _col_idx('盈利能力_趋势')
        growth_slope_ci = _col_idx('成长能力_斜率')
        growth_trend_ci = _col_idx('成长能力_趋势')
        combined_ci = _col_idx('综合趋势')
        ret7m_ci = _col_idx('7个月后涨跌幅')
        price7m_ci = _col_idx('7个月后价格')
        final_ci = _col_idx('最终结论')
        # 子场景列
        sub_cols = {
            'sub_market_index': _col_idx('市场指数'),
            'sub_industry_pe': _col_idx('行业PE'),
            'sub_stock_pe': _col_idx('个股PE'),
            'sub_dcf': _col_idx('DCF估值'),
            'sub_adj_pe': _col_idx('修正PE估值'),
            'sub_param_build': _col_idx('参数构造'),
            'sub_monte_carlo': _col_idx('蒙特卡洛'),
            'sub_reverse_calc': _col_idx('反向推算'),
        }

        saved = 0
        for row_idx in range(2, ws.max_row + 1):
            stock_code = str(ws.cell(row_idx, code_col_idx).value or '').strip()
            if not stock_code:
                continue

            def _val(ci):
                if ci is None:
                    return None
                v = ws.cell(row_idx, ci).value
                return v

            def _num(ci):
                if ci is None:
                    return None
                v = ws.cell(row_idx, ci).value
                if v is None:
                    return None
                try:
                    return float(str(v).replace('%', '').replace('+', ''))
                except (ValueError, TypeError):
                    return v

            def _sub_val(ci):
                """子场景: ✓→1, 其他→0"""
                if ci is None:
                    return None
                v = str(ws.cell(row_idx, ci).value or '')
                return 1 if '✓' in v else 0

            def _pct_decimal(ci):
                """溢价率百分比 → 小数: '-20.00%'→-0.2；无效('-'/空)→None。
                与 backfill_evaluations.py 及 export_features 的 Excel 加载器(/100)保持一致。"""
                if ci is None:
                    return None
                v = _num(ci)
                if not isinstance(v, (int, float)):
                    return None
                return v / 100.0

            data = {
                'stock_code': stock_code,
                'stock_name': str(_val(name_col_idx) or ''),
                'batch_id': f'{date_tag}_{input_basename}',
                'issue_date': str(_val(quote_date_ci) or '') if quote_date_ci else None,
                'issue_date_price': _num(quote_price_ci),
                'valid_thresholds': _num(thresholds_ci),
                'premium_min': _pct_decimal(prem_min_ci),
                'premium_max': _pct_decimal(prem_max_ci),
                'decision': str(_val(decision_ci) or ''),
                'industry_l1': str(_val(l1_ci) or '') if l1_ci else None,
                'industry_l2': str(_val(l2_ci) or '') if l2_ci else None,
                'industry_l3': str(_val(l3_ci) or '') if l3_ci else None,
                'total_slope': _num(total_slope_ci),
                'total_trend': str(_val(total_trend_ci) or '') if total_trend_ci else None,
                'profit_slope': _num(profit_slope_ci),
                'profit_trend': str(_val(profit_trend_ci) or '') if profit_trend_ci else None,
                'growth_slope': _num(growth_slope_ci),
                'growth_trend': str(_val(growth_trend_ci) or '') if growth_trend_ci else None,
                'combined_trend': str(_val(combined_ci) or '') if combined_ci else None,
                'return_7m': _num(ret7m_ci),
                'price_7m': _num(price7m_ci),
                'final_conclusion': str(_val(final_ci) or '') if final_ci else None,
            }
            # 子场景
            for key, ci in sub_cols.items():
                data[key] = _sub_val(ci)

            db.save_placement_evaluation(data)
            saved += 1

        print(f'  ✅ 落库完成: {saved}条记录 → placement_evaluation [耗时 {time.time()-t_db:.1f}s]')
    except Exception as e:
        print(f'  ⚠️ 落库失败(不影响Excel输出): {e}')

    step3_elapsed = time.time() - t_step3
    total = ws.max_row - 1
    print()
    print(f'  最终结论: {pass_count}/{total} 通过 [耗时 {step3_elapsed:.1f}s]')

    # ── Step 4: ML模型预测盈利概率 ──
    if not os.path.exists(PREDICT_SCRIPT):
        print(f'\n⚠️ ML预测脚本不存在: {PREDICT_SCRIPT}，跳过 Step 4')
    else:
        print()
        print('=' * 60)
        print('  Step 4/4: ML模型预测盈利概率（LightGBM + 逻辑回归）')
        print('=' * 60)
        step4_cmd = [
            PYTHON, PREDICT_SCRIPT,
            scored_output,
        ]
        if not run_step('Step 4/4: ML盈利概率预测', step4_cmd):
            print('⚠️ ML预测失败，不影响前面步骤的结果')

    print()
    total_elapsed = time.time() - t_total
    print('=' * 60)
    print(f'🎉 全部完成! [总耗时 {total_elapsed:.1f}s]')
    print(f'   筛选结果: {screening_output}')
    print(f'   最终结果: {scored_output}')
    print('=' * 60)


if __name__ == '__main__':
    main()
