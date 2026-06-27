# ====== batch_financial_score (score) ======

import sys
import os
import time
import threading
import logging

import pandas as pd
import numpy as np
import tushare as ts
import openpyxl
import warnings

warnings.filterwarnings('ignore', category=FutureWarning)

_EFAES_ROOT = os.path.expanduser('~/github/EFAES')
sys.path.insert(0, _EFAES_ROOT)
# PKG 根(price_maintenance_risk_analysis/): utils.db_manager 所在, 供 bulk_load_* 复用
_PKG_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _PKG_ROOT not in sys.path:
    sys.path.insert(0, _PKG_ROOT)

from src.web.config import DB_CONFIG, TUSHARE_TOKEN
from src.core.tushare_data_fetcher import TushareDataFetcher
from src.core.industry_threshold_calculator import IndustryThresholdCalculator
from src.core.calculate_financial_score import FinancialScoringCard
from src.core.calculate_financial_ratios import calculate_financial_ratios

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

CURRENT_YEAR = 2026
SCORE_YEARS = list(range(CURRENT_YEAR - 5, CURRENT_YEAR))  # 默认，会被每只股票的报价日覆盖
RELATIVE_YEAR_LABELS = ['T-4', 'T-3', 'T-2', 'T-1', 'T']  # 相对年份表头

# 行业阈值缓存，同行业不重复计算
_industry_thresholds_cache = {}  # {(industry, year): thresholds}
_industry_raw_cache = {}  # {industry: (df_with_stats, fetch_end_year)} 原始数据缓存

# 行业查询节流锁：多线程并发查 index_member_all 会打爆 tushare 限频，
# 限频后接口返回空 → 行业缺失 → 整行评分被跳过。用此锁把行业查询串行化 + 每次间隔，
# 配合下面的重试，避免限频导致的评分缺失。
_industry_query_lock = threading.Lock()

# 二/一级行业回退阈值缓存：冷门三级行业样本不足时回退到二级/一级，同行业只算一次
_fallback_thresholds_cache = {}  # {(industry_name, year): thresholds}

# ──────────────────────────────────────────────────────────────────
# ③-d 加速: 比率 + 行业映射 改读本地已落库表(替逐股 tushare); 阈值加持久缓存。
# 字节级等同原路径:
#   - financial_indicators 由 backfill_financial_indicators(③-c) 用【同一 calculate_financial_ratios】落库
#   - industry_data 是 ③-a 的申万映射(同 SW 口径)
#   - 阈值仍走原 EFAES 算法(tushare fina_indicator 快捷值分布, 与评分源一致), 仅加持久缓存免重启重算
# ──────────────────────────────────────────────────────────────────
_FI_RATIO_COLS = [
    'current_ratio', 'quick_ratio', 'inv_turn', 'ar_turn', 'ca_turn', 'assets_turn',
    'roa', 'npta', 'roe', 'roe_dt', 'netprofit_margin', 'grossprofit_margin',
    'debt_to_assets', 'int_to_talcap', 'debt_to_eqt', 'ebit_to_interest',
    'cash_to_liqdebt', 'cash_to_liqdebt_withinterest', 'rd_exp_ratio',
    'op_yoy', 'ebt_yoy', 'netprofit_yoy', 'dt_netprofit_yoy', 'roe_yoy',
    'tr_yoy', 'or_yoy', 'equity_yoy',
]
_RATIOS_CACHE = {}     # {(ts_code, year): {ratio: val}}   批量预取一次, 评分线程只读
_INDUSTRY_CACHE = {}   # {ts_code: {l1_name,l2_name,l3_name}}  本地 industry_data 预取
_FI_DF = pd.DataFrame()        # financial_indicators 全表(阈值分位用): stock_code+report_year+27比率
_INDUSTRY_MEMBERS = {}         # {l3/l2/l1_name: [stock_codes]}  本地反查, 替 get_industry_stocks

_THRESHOLD_CACHE_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    'ml_training', 'data', 'threshold_cache.json')
_persistent_threshold_cache = {}   # {f"{industry}|{year}": thresholds}   跨进程持久
_threshold_save_lock = threading.Lock()


def bulk_load_ratios(stock_codes):
    """一次性把全 universe 的 (stock,year)→27比率 读进内存。
    financial_indicators 由 ③-c 用同一 calculate_financial_ratios 落库 → 读它字节级等同原取数。"""
    import pymysql
    from utils.db_manager import ValuationDB
    _RATIOS_CACHE.clear()
    codes = [str(c) for c in stock_codes if c]
    cfg = ValuationDB.MYSQL_CONFIG
    conn = pymysql.connect(host=cfg['host'], port=cfg['port'], user=cfg['user'],
                           password=cfg['password'], database=cfg['database'], charset=cfg['charset'])
    cols = ','.join(_FI_RATIO_COLS)
    BATCH = 500
    frames = []
    try:
        for i in range(0, len(codes), BATCH):
            b = codes[i:i + BATCH]
            ph = ','.join(['%s'] * len(b))
            frames.append(pd.read_sql(
                f"SELECT stock_code,report_year,{cols} FROM financial_indicators WHERE stock_code IN ({ph})",
                conn, params=b))
    finally:
        conn.close()
    df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    global _FI_DF
    _FI_DF = df.copy()   # 阈值分位用(全表, stock_code+report_year+27比率)
    hit_stocks = set()
    for _, row in df.iterrows():
        _RATIOS_CACHE[(str(row['stock_code']), int(row['report_year']))] = {
            c: (row[c] if pd.notna(row[c]) else None) for c in _FI_RATIO_COLS}
        hit_stocks.add(str(row['stock_code']))
    logger.info(f"比率缓存(financial_indicators): {len(_RATIOS_CACHE)} 条, {len(hit_stocks)}/{len(codes)} 股命中")


def bulk_load_industry():
    """从本地 industry_data(③-a 申万映射) 预取 ts_code→l1/l2/l3_name, 替 5302 次 index_member_all。
    每股取最新 analysis_date; 无 l3 的不存(评分按缺 l3 回退逐股查, 同原逻辑)。"""
    import pymysql
    from utils.db_manager import ValuationDB
    _INDUSTRY_CACHE.clear()
    cfg = ValuationDB.MYSQL_CONFIG
    conn = pymysql.connect(host=cfg['host'], port=cfg['port'], user=cfg['user'],
                           password=cfg['password'], database=cfg['database'], charset=cfg['charset'])
    try:
        df = pd.read_sql(
            "SELECT stock_code,sw_l1_name,sw_l2_name,sw_l3_name,analysis_date "
            "FROM industry_data WHERE sw_l3_name IS NOT NULL AND sw_l3_name<>''",
            conn)
    finally:
        conn.close()
    if df.empty:
        logger.warning("industry_data 为空(l3), 行业查询回退逐股 index_member_all")
        return
    df = df.sort_values('analysis_date', ascending=False).drop_duplicates(subset=['stock_code'], keep='first')
    for _, row in df.iterrows():
        _INDUSTRY_CACHE[str(row['stock_code'])] = {
            'l1_name': row.get('sw_l1_name') or '',
            'l2_name': row.get('sw_l2_name') or '',
            'l3_name': row.get('sw_l3_name') or '',
        }
    # 反查 {l3/l2/l1_name: [codes]} 供本地阈值(替 get_industry_stocks 的 tushare 调用)
    global _INDUSTRY_MEMBERS
    _INDUSTRY_MEMBERS = {}
    for lvl in ['sw_l3_name', 'sw_l2_name', 'sw_l1_name']:
        for name, g in df.dropna(subset=[lvl]).groupby(lvl):
            _INDUSTRY_MEMBERS.setdefault(str(name), []).extend(g['stock_code'].astype(str).tolist())
    n_l3 = df['sw_l3_name'].dropna().nunique() if 'sw_l3_name' in df.columns else 0
    logger.info(f"行业缓存(industry_data 本地): {len(_INDUSTRY_CACHE)} 股; 成员反查 {len(_INDUSTRY_MEMBERS)} 桶(L3={n_l3})")


def _load_persistent_thresholds():
    """幂等载入持久阈值缓存(跨进程/重启复用)。"""
    global _persistent_threshold_cache
    if _persistent_threshold_cache:
        return
    if os.path.exists(_THRESHOLD_CACHE_FILE):
        try:
            import json
            with open(_THRESHOLD_CACHE_FILE, 'r') as f:
                _persistent_threshold_cache = json.load(f)
            logger.info(f"持久阈值缓存: 载入 {len(_persistent_threshold_cache)} 条")
        except Exception as e:
            logger.warning(f"载入持久阈值缓存失败: {e}")


def _save_persistent_threshold(key, thresholds):
    """阈值算好后落盘(算法不变, 仅缓存; 线程安全)。"""
    _persistent_threshold_cache[key] = thresholds
    try:
        import json
        with _threshold_save_lock:
            os.makedirs(os.path.dirname(_THRESHOLD_CACHE_FILE), exist_ok=True)
            with open(_THRESHOLD_CACHE_FILE, 'w') as f:
                json.dump(_persistent_threshold_cache, f)
    except Exception as e:
        logger.debug(f"写持久阈值缓存失败: {e}")


def get_company_industry(pro, ts_code: str, max_retries: int = 3) -> dict:
    """通过 Tushare 反查申万三级行业，返回 {l1_name, l2_name, l3_name}

    带节流+重试：index_member_all 接口在并发查询时易触发限频返回空，
    用全局锁串行化调用 + 每次间隔，并在返回空/异常时退避重试，避免限频导致评分缺失。
    """
    result = {'l1_name': '', 'l2_name': '', 'l3_name': ''}
    for attempt in range(max_retries):
        try:
            # index_member_all 官方支持并发(单次≤2000行), 直接调; 退避重试兜底瞬时限频
            df = pro.index_member_all(ts_code=ts_code)
            if df is not None and not df.empty:
                row = df.sort_values('in_date', ascending=False).iloc[0]
                l1 = row.get('l1_name', '')
                l2 = row.get('l2_name', '')
                l3 = row.get('l3_name', '')
                if l3:
                    logger.info(f"{ts_code} 申万行业: {l1} > {l2} > {l3}")
                    return {'l1_name': l1, 'l2_name': l2, 'l3_name': l3}
                # df 非空但 l3 空：数据本身无三级分类，不重试
                return {'l1_name': l1, 'l2_name': l2, 'l3_name': l3}
            # df 为空：疑似限频，退避后重试
            logger.warning(f"{ts_code} 行业查询返回空(尝试 {attempt + 1}/{max_retries})，疑似限频")
        except Exception as e:
            logger.warning(f"查询 {ts_code} 申万行业失败(尝试 {attempt + 1}/{max_retries}): {e}")
        if attempt < max_retries - 1:
            time.sleep(1.0 * (attempt + 1))  # 退避：1s, 2s

    logger.warning(f"{ts_code} 行业查询 {max_retries} 次均失败/为空，评分将被跳过")
    return result


def _get_industry_stocks_by_level(pro, industry_name, level='L2'):
    """获取申万 L2/L1 行业的成分股列表（阈值兜底用）。

    EFAES 的 get_industry_stocks 只认三级行业名（硬编码 level='L3'）；
    本函数补充 L2/L1，用于三级行业样本不足时回退。返回 ts_code 列表，失败返回 []。
    """
    try:
        df = pro.index_classify(level=level, src='SW2021')
        if df is None or df.empty:
            return []
        matched = df[df['industry_name'] == industry_name]
        if matched.empty:
            matched = df[df['industry_name'].str.contains(industry_name, na=False)]
        if matched.empty:
            logger.warning(f"申万{level} 行业 '{industry_name}' 未匹配到指数")
            return []
        index_code = matched.iloc[0]['index_code']
        logger.info(f"申万{level}匹配: '{industry_name}' -> {index_code}")
        members = pro.index_member(index_code=index_code, is_new='Y')
        if members is None or members.empty:
            return []
        return members['con_code'].unique().tolist()
    except Exception as e:
        logger.warning(f"申万{level}行业 '{industry_name}' 成分股查询失败: {e}")
        return []


def fetch_financial_data(fetcher: TushareDataFetcher, ts_code: str, score_years=None, max_retries: int = 3):
    """获取公司三张报表数据，范围覆盖score_years（需前一年做同比）。

    带重试：fina_indicator/财报接口在并发下易触发 tushare 限频(500次/分钟)返回空，
    空时退避重试，避免限频导致个股三表为空、评分被跳过。
    """
    years = score_years if score_years else SCORE_YEARS
    start_date = f'{min(years) - 1}0101'
    end_date = f'{max(years)}1231'

    def _try(fetch_fn, label):
        for attempt in range(max_retries):
            try:
                df = fetch_fn(ts_code, start_date, end_date)
                if df is not None and not df.empty:
                    return df
            except Exception as e:
                logger.warning(f"{ts_code} {label}获取异常(尝试 {attempt + 1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(1.0 * (attempt + 1))  # 退避：1s, 2s
        return pd.DataFrame()

    bs = _try(fetcher.fetch_balance_sheet, '资产负债表')
    inc = _try(fetcher.fetch_income_statement, '利润表')
    cf = _try(fetcher.fetch_cash_flow_statement, '现金流量表')

    for df in [bs, inc, cf]:
        if not df.empty:
            df['report_year'] = df['end_date'].apply(lambda x: int(str(x)[:4]) if pd.notna(x) else None)
            company_name = fetcher._get_company_name(ts_code) or ts_code
            df['company_name'] = company_name

    return bs, inc, cf


def get_industry_thresholds(threshold_calc: IndustryThresholdCalculator, industry_name: str,
                            target_year: int = None) -> dict:
    """获取行业阈值（按年份缓存），优先用target_year，不足则往前找。

    Args:
        target_year: 目标年份（报价日年份），阈值反映该年行业分布。
                     None则用SCORE_YEARS[-1]。
    """
    global _industry_thresholds_cache

    # 按行业+年份组合缓存（不同年份行业阈值不同）
    cache_key = (industry_name, target_year)
    if cache_key in _industry_thresholds_cache:
        logger.info(f"使用缓存的 '{industry_name}' {target_year}年 行业阈值")
        return _industry_thresholds_cache[cache_key]

    # 持久缓存(跨进程复用, 算法不变): 命中则回放, 免 tushare 重算
    _load_persistent_thresholds()
    pkey = f"{industry_name}|{target_year}"
    if pkey in _persistent_threshold_cache:
        cached = _persistent_threshold_cache[pkey]
        if cached:
            _industry_thresholds_cache[cache_key] = cached
            logger.info(f"复用持久缓存 '{industry_name}' {target_year}年 行业阈值")
            return cached

    # 确定尝试的年份：从target_year往前找（最多回退3年）
    base_year = target_year if target_year else SCORE_YEARS[-1]
    try_years = [base_year - i for i in range(4)]  # target_year, T-1, T-2, T-3

    # 原始数据缓存：窄窗覆盖 try_years[base_year..base_year-3]。
    # ⚠ 必须【窄窗】(end=base_year): tushare fina_indicator 每次最多100行取最新,
    # 宽窗(→CURRENT_YEAR)会把早年挤掉 → create_thresholds 找不到 target_year → 退化。
    fetch_start_year = base_year - 4
    fetch_end_year = base_year
    df_with_stats = None
    if industry_name in _industry_raw_cache:
        cached_stats, cached_start, cached_end = _industry_raw_cache[industry_name]
        # 缓存数据需同时覆盖start和end
        if cached_start <= fetch_start_year and cached_end >= base_year:
            df_with_stats = cached_stats
            logger.info(f"复用 '{industry_name}' 已缓存的原始数据（{cached_start}~{cached_end}年）")

    if df_with_stats is None:
        # ── 本地阈值路径(替 tushare 逐成员 fina_indicator, 限频卡死):
        # 从 financial_indicators(③-c, 同 calculate_financial_ratios 年报口径)取行业成员比率,
        # add_industry_statistics 注入分位行, create_thresholds_from_dataframe 原算法不变。
        # 成员从 industry_data(③-a 本地)反查; 年报口径 ≈ 快捷值年报。零 tushare。 ──
        members = _INDUSTRY_MEMBERS.get(industry_name, [])
        if not members:
            logger.warning(f"行业 '{industry_name}' 本地无成分股(industry_data), 跳过")
            _industry_thresholds_cache[cache_key] = {}
            return {}
        sub = _FI_DF[_FI_DF['stock_code'].isin(members)].copy()
        if sub.empty:
            logger.warning(f"行业 '{industry_name}' 成员在 financial_indicators 无数据, 跳过")
            _industry_thresholds_cache[cache_key] = {}
            return {}
        sub = sub.rename(columns={'stock_code': 'ts_code', 'report_year': 'year'})
        sub['company_name'] = sub['ts_code'].astype(str)
        try:
            from src.core.tushare_tools import add_industry_statistics
            df_with_stats = add_industry_statistics(sub)
            _industry_raw_cache[industry_name] = (df_with_stats, 2006, 2026)   # 本地表覆盖全年份
            logger.info(f"✅ '{industry_name}' 本地阈值数据({len(members)}成员×{len(sub)}行)")
        except Exception as e:
            logger.error(f"行业 '{industry_name}' 本地阈值构建失败: {e}")
            _industry_thresholds_cache[cache_key] = {}
            return {}

    # 从已缓存的原始数据计算各年阈值
    thresholds = {}
    for try_year in try_years:
        try:
            result = threshold_calc.create_thresholds_from_dataframe(df_with_stats, try_year, industry_name)
            if not result or 'thresholds' not in result:
                continue
            t = result['thresholds']
            valid = sum(1 for v in t.values() if v is not None)
            if valid >= 10:
                logger.info(f"行业 '{industry_name}' 使用 {try_year} 年阈值（有效指标: {valid}）")
                thresholds = t
                break
        except Exception as e:
            logger.warning(f"从缓存数据计算 {try_year} 年阈值失败: {e}")
            continue

    if not thresholds:
        logger.error(f"计算行业 '{industry_name}' 的阈值失败")
        return {}

    clean_thresholds = {k: v for k, v in thresholds.items() if v is not None}
    _industry_thresholds_cache[cache_key] = clean_thresholds
    _save_persistent_threshold(f"{industry_name}|{target_year}", clean_thresholds)
    return clean_thresholds


def _calc_thresholds_for_codes(threshold_calc, ts_codes, industry_name, target_year=None):
    """用外部成分股 ts_codes 计算行业阈值（二/一级 fallback 用），带年份回退与内存缓存。

    复用 EFAES calculate_custom_industry_thresholds（接收外部 ts_codes，绕开只认三级的
    get_industry_stocks）。冷门三级行业样本不足时，用其所属二/一级的成分股算阈值兜底。
    """
    base_year = target_year if target_year else SCORE_YEARS[-1]
    for try_year in [base_year - i for i in range(4)]:
        cache_key = (industry_name, try_year)
        if cache_key in _fallback_thresholds_cache:
            logger.info(f"复用回退行业 '{industry_name}' {try_year}年阈值(缓存)")
            return _fallback_thresholds_cache[cache_key]
        try:
            result = threshold_calc.calculate_custom_industry_thresholds(
                industry_name, ts_codes, target_year=try_year, save_to_db=False)
            if not result or 'thresholds' not in result:
                continue
            t = result.get('thresholds') or {}
            valid = sum(1 for v in t.values() if v is not None)
            if valid >= 10:
                clean = {k: v for k, v in t.items() if v is not None}
                logger.info(f"回退行业 '{industry_name}' 使用 {try_year} 年阈值（有效指标 {valid}/{len(t)}）")
                _fallback_thresholds_cache[cache_key] = clean
                return clean
        except Exception as e:
            logger.warning(f"回退行业 '{industry_name}' {try_year}年阈值计算失败: {e}")
            continue
    return {}


def _save_financial_indicators(stock_code, report_year, ratios):
    """将计算好的财务指标保存到 investment_valuation.financial_indicators 表"""
    import pymysql

    # 指标字段列表(与表结构一一对应)
    indicator_fields = [
        'current_ratio', 'quick_ratio', 'inv_turn', 'ar_turn', 'ca_turn', 'assets_turn',
        'roa', 'npta', 'roe', 'roe_dt', 'netprofit_margin', 'grossprofit_margin',
        'debt_to_assets', 'int_to_talcap', 'debt_to_eqt', 'ebit_to_interest',
        'cash_to_liqdebt', 'cash_to_liqdebt_withinterest', 'rd_exp_ratio',
        'op_yoy', 'ebt_yoy', 'netprofit_yoy', 'dt_netprofit_yoy', 'roe_yoy',
        'tr_yoy', 'or_yoy', 'equity_yoy',
    ]

    # 构建 VALUES 部分
    values = [stock_code, report_year]
    placeholders = ['%s', '%s']
    update_parts = []
    for f in indicator_fields:
        values.append(ratios.get(f))
        placeholders.append('%s')
        update_parts.append(f'{f}=VALUES({f})')

    values.append(1 if ratios.get('used_average_values') else 0)
    placeholders.append('%s')
    update_parts.append('used_average_values=VALUES(used_average_values)')

    values.append(ratios.get('valid_indicator_count', 0))
    placeholders.append('%s')
    update_parts.append('valid_indicator_count=VALUES(valid_indicator_count)')

    conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='',
                           database='investment_valuation', charset='utf8mb4')
    try:
        with conn.cursor() as cur:
            cols = ', '.join(['stock_code', 'report_year'] + indicator_fields +
                             ['used_average_values', 'valid_indicator_count'])
            ph = ', '.join(placeholders)
            sql = f"INSERT INTO financial_indicators ({cols}) VALUES ({ph}) ON DUPLICATE KEY UPDATE {', '.join(update_parts)}"
            cur.execute(sql, values)
            conn.commit()
    finally:
        conn.close()


def _save_annual_scores_to_db(stock_code, results, stock_years):
    """将年度财务评分保存到 investment_valuation.company_annual_scores"""
    import sys as _sys
    _proj_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    if _proj_root not in _sys.path:
        _sys.path.insert(0, _proj_root)
    from utils.db_manager import ValuationDB
    db = ValuationDB()

    scores_by_year = {}
    for year in stock_years:
        if year in results and isinstance(results[year], dict):
            r = results[year]
            scores_by_year[year] = {
                'stock_name': None,
                'total_score': r.get('总分'),
                'rating': r.get('评级'),
                'profitability': r.get('盈利能力'),
                'growth': r.get('成长能力'),
                'operating': r.get('运营能力'),
                'solvency': r.get('偿债能力'),
            }

    # 补充行业信息
    industry_info = results.get('行业') or {}
    for year_data in scores_by_year.values():
        year_data['industry_l1'] = industry_info.get('l1_name')
        year_data['industry_l2'] = industry_info.get('l2_name')
        year_data['industry_l3'] = industry_info.get('l3_name')

    if scores_by_year:
        db.save_annual_scores(stock_code, scores_by_year)


def score_company(fetcher: TushareDataFetcher, threshold_calc: IndustryThresholdCalculator,
                  ts_code: str, company_name: str, score_years=None) -> dict:
    """计算单个公司近5年的财务评分

    Args:
        score_years: 评分年份列表（从报价日回溯），None则用全局SCORE_YEARS
    """
    pro = fetcher.pro
    years_to_score = score_years if score_years else SCORE_YEARS
    logger.info(f"{'='*60}")
    logger.info(f"开始处理: {ts_code} {company_name}")
    logger.info(f"{'='*60}")

    # 1. 获取行业: 优先本地预取(industry_data, ③-a 同 SW 口径), 缺则回退逐股 index_member_all
    industry_info = _INDUSTRY_CACHE.get(ts_code) or get_company_industry(pro, ts_code)
    if not industry_info.get('l3_name'):
        logger.error(f"无法确定 {ts_code} 的三级行业，跳过")
        return {}

    # 2. 获取行业阈值：三级 → 二级 → 一级 回退（冷门三级行业样本不足时兜底）
    # ⚠ target_year=该股报价日年(T): 行业阈值逐年不同, 必须按股的 T 年取; 复用只在
    # (行业, T年) 相同时命中。原代码不传 target_year → 全部用 SCORE_YEARS[-1](首行年),
    # 混合年批次会把 2010 阈值套到 2015 股上(年份错配)。
    industry_name = industry_info['l3_name']
    t_year = int(years_to_score[-1])
    thresholds = get_industry_thresholds(threshold_calc, industry_name, target_year=t_year)
    if not thresholds:
        for level_name, key, sw_level in (('二级', 'l2_name', 'L2'), ('一级', 'l1_name', 'L1')):
            fb_name = industry_info.get(key, '')
            if not fb_name:
                continue
            fb_codes = _get_industry_stocks_by_level(pro, fb_name, level=sw_level)
            if not fb_codes:
                continue
            thresholds = _calc_thresholds_for_codes(threshold_calc, fb_codes, fb_name, target_year=t_year)
            if thresholds:
                logger.info(f"{ts_code} 三级行业 '{industry_name}' 阈值样本不足，回退{level_name}行业 '{fb_name}'({len(fb_codes)}只成分股)")
                industry_name = fb_name
                break
    if not thresholds:
        logger.error(f"{ts_code} 三/二/一级行业阈值均计算失败，跳过")
        return {}

    logger.info(f"行业阈值加载完成（{industry_name}），共 {len(thresholds)} 个指标")
    valid_count = sum(1 for v in thresholds.values() if v is not None)
    logger.info(f"有效阈值: {valid_count}/{len(thresholds)}")

    # 3. 财务比率: 优先读 financial_indicators 内存缓存(③-c 同 calculate_financial_ratios 落库, 字节级等同),
    #    缺失年份才回退 tushare 取三表现算(覆盖4缺股/年份gap), 算完补落库
    ratios_by_year = {}
    missing_years = []
    for y in years_to_score:
        r = _RATIOS_CACHE.get((ts_code, y))
        if r:
            ratios_by_year[y] = r
        else:
            missing_years.append(y)
    if missing_years:
        bs, inc, cf = fetch_financial_data(fetcher, ts_code, score_years=missing_years)
        if not (bs.empty or inc.empty):
            for year in missing_years:
                bs_year = bs[bs['report_year'] == year]
                inc_year = inc[inc['report_year'] == year]
                if bs_year.empty or inc_year.empty:
                    continue
                r = calculate_financial_ratios(bs, inc, cf, year)
                if r:
                    ratios_by_year[year] = r
                    try:
                        _save_financial_indicators(ts_code, year, r)
                    except Exception as e:
                        logger.debug(f"{ts_code} {year}年 指标落库失败(不影响评分): {e}")
    if not ratios_by_year:
        logger.error(f"{ts_code} 无可用财务比率(DB+tushare均空)，跳过")
        return {}

    # 4. 逐年评分
    results = {}
    scorer = FinancialScoringCard(DB_CONFIG)
    scorer.set_industry_thresholds(thresholds)

    for year in years_to_score:
        try:
            ratios = ratios_by_year.get(year)
            if not ratios:
                logger.warning(f"{ts_code} {year}年 比率缺失，跳过")
                continue

            scorer.load_company_data(ratios, company_name)
            score_result = scorer.calculate_score()

            total_score = score_result['总得分']
            rating = scorer.get_rating(total_score)
            dim_scores = score_result['维度得分']

            results[year] = {
                '总分': round(total_score, 2),
                '评级': rating,
                '运营能力': round(dim_scores.get('运营能力', 0), 2),
                '盈利能力': round(dim_scores.get('盈利能力', 0), 2),
                '偿债能力': round(dim_scores.get('偿债能力', 0), 2),
                '成长能力': round(dim_scores.get('成长能力', 0), 2),
            }
            logger.info(f"  {year}年: 总分={total_score:.2f} 评级={rating}")

        except Exception as e:
            logger.warning(f"{ts_code} {year}年 评分计算失败: {e}")
            import traceback
            traceback.print_exc()
            continue

    results['行业'] = industry_info
    return results


def main_score():
    import argparse
    parser = argparse.ArgumentParser(description='批量财务评分（支持行级增量续跑）')
    parser.add_argument('excel_path', help='输入Excel（screening结果）；若对应 _scored 已存在则续传补算')
    parser.add_argument('--force', action='store_true', help='强制全量重算（忽略已有 _scored，从头评分）')
    parser.add_argument('--years-wide', type=int, default=5,
                        help='评分年份窗口宽度(默认5=报价年及前4年, 与既有表头一致); '
                             '回测全历史用 16(报价年及前15年, 2010~2025)——宽窗强制从原 Excel 重建, 不复用 5 年 _scored')
    args = parser.parse_args()
    years_wide = args.years_wide

    excel_path = args.excel_path
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        sys.exit(1)

    # 续传模式：若 _scored 已存在且非 --force，则读它（保留已评分行，只补失败行）
    output_path = excel_path.replace('.xlsx', '_scored.xlsx')
    resume_mode = (not args.force) and os.path.exists(output_path)
    # 宽窗(回测全历史)强制从原 Excel 重建: 既有 _scored 是 5 年表头(T-4..T),
    # 与 16 年窗口列数不符, 续写会错位 → 必须重建。
    if years_wide != 5 and resume_mode:
        resume_mode = False
        logger.info(f"--years-wide {years_wide}: 从原 Excel 重建(忽略既有 5 年 _scored 表头)")
    src_path = output_path if resume_mode else excel_path
    logger.info(f"读取 Excel: {src_path}" + ("（行级续传：仅补算评分失败的行）" if resume_mode else "（全量评分）"))

    # 初始化
    pro = ts.pro_api(TUSHARE_TOKEN)
    fetcher = TushareDataFetcher(TUSHARE_TOKEN, DB_CONFIG)
    threshold_calc = IndustryThresholdCalculator(TUSHARE_TOKEN, DB_CONFIG)

    # 读取 Excel
    wb = openpyxl.load_workbook(src_path)
    ws = wb.active

    # 找到股票代码列
    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    code_col = header_row.index('股票代码') + 1
    name_col = header_row.index('股票简称') + 1

    # 读取报价日列（如有），按报价日年份回溯确定评分年份
    quote_date_col = None
    if '报价日' in header_row:
        quote_date_col = header_row.index('报价日') + 1
    # 取第一只有报价日的股票确定评分年份范围
    base_year = CURRENT_YEAR
    if quote_date_col:
        for row_idx in range(2, ws.max_row + 1):
            qd = ws.cell(row_idx, quote_date_col).value
            if qd and str(qd).strip() and len(str(qd).strip()) >= 4:
                try:
                    base_year = int(str(qd).strip()[:4])
                    logger.info(f"检测到报价日 {qd}，评分年份按报价日({base_year}年)回溯")
                    break
                except ValueError:
                    pass
    score_years = list(range(base_year - (years_wide - 1), base_year + 1))  # 报价年及前 (years_wide-1) 年
    logger.info(f"评分年份: {score_years} (宽度 {years_wide})")
    # 更新全局SCORE_YEARS，使所有函数(get_industry_thresholds/score_company)同步使用
    global SCORE_YEARS
    SCORE_YEARS = score_years

    # 表头定位/初始化（幂等：已有评分列则复用列索引，否则新增——避免续跑重复加一整套列）
    fields = ['总分', '评级', '盈利能力', '成长能力']
    check_fields_trend = ['总分', '盈利能力', '成长能力']

    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    if '三级行业' in header_row and '总分_T-4' in header_row and '总分_斜率' in header_row:
        # 续传：复用已有评分列（列顺序与首次写入一致）
        start_col = header_row.index('一级行业') + 1
        data_start_col = header_row.index('总分_T-4') + 1
        trend_start = header_row.index('总分_斜率') + 1
        logger.info("检测到已有评分列，复用列位置（续传模式）")
    else:
        # 首次：追加表头
        start_col = ws.max_column + 1
        ws.cell(1, start_col, '一级行业')
        ws.cell(1, start_col + 1, '二级行业')
        ws.cell(1, start_col + 2, '三级行业')
        data_start_col = start_col + 3

        rel_labels = ([f'T-{years_wide - 1 - i}' for i in range(years_wide)])
        if rel_labels:
            rel_labels[-1] = 'T'  # 末位 = 报价年(宽度5时退化为 ['T-4','T-3','T-2','T-1','T'])
        headers = [f'{field}_{label}' for field in fields for label in rel_labels]
        for i, h in enumerate(headers):
            ws.cell(1, data_start_col + i, h)

        trend_start = data_start_col + len(headers)
        trend_headers = []
        for field in check_fields_trend:
            trend_headers.append(f'{field}_斜率')
            trend_headers.append(f'{field}_趋势')
        trend_headers.append('综合趋势')
        for i, h in enumerate(trend_headers):
            ws.cell(1, trend_start + i, h)

    # 刷新表头（首次模式下含新加列；续传模式下不变）
    header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]

    # ===== 并发评分（多线程，API调用为主，适合I/O并行）=====
    import threading
    from concurrent.futures import ThreadPoolExecutor, as_completed

    _write_lock = threading.Lock()
    _done_count = [0]
    # _total 在下方确定待补算行（row_indices）后设置

    def _score_one(row_idx):
        """评分单只股票（线程内），返回(row_idx, results, score_years)"""
        ts_code = ws.cell(row_idx, code_col).value
        company_name = ws.cell(row_idx, name_col).value
        if not ts_code:
            return row_idx, None, None
        ts_code = str(ts_code).strip()
        # 读取该股票的报价日，计算各自回溯的评分年份
        stock_years = SCORE_YEARS  # 默认
        if quote_date_col:
            qd = ws.cell(row_idx, quote_date_col).value
            if qd and str(qd).strip() and len(str(qd).strip()) >= 4:
                try:
                    base_year = int(str(qd).strip()[:4])
                    stock_years = list(range(base_year - (years_wide - 1), base_year + 1))
                except ValueError:
                    pass
        results = score_company(fetcher, threshold_calc, ts_code, company_name, score_years=stock_years)
        return row_idx, results, stock_years

    def _write_results(row_idx, results, stock_years):
        """将评分结果写入Excel（线程安全，按相对年份顺序写入）"""
        if results is None:
            return
        # 写入三级行业信息
        if results and results.get('行业'):
            info = results['行业']
            ws.cell(row_idx, start_col, info.get('l1_name', ''))
            ws.cell(row_idx, start_col + 1, info.get('l2_name', ''))
            ws.cell(row_idx, start_col + 2, info.get('l3_name', ''))
            results.pop('行业')

        # 按stock_years顺序写入（对应表头T-4,T-3,...,T）
        col_offset = data_start_col
        for field in fields:
            for year in stock_years:
                if year in results:
                    ws.cell(row_idx, col_offset, results[year].get(field, 'N/A'))
                else:
                    ws.cell(row_idx, col_offset, 'N/A')
                col_offset += 1

        if results:
            check_fields_t = ['总分', '盈利能力', '成长能力']
            thresholds = {'总分': -2, '盈利能力': -1, '成长能力': -1}
            fail_count = 0
            for field in check_fields_t:
                scores = [results[y][field] for y in stock_years if y in results and isinstance(results[y].get(field), (int, float))]
                slope = np.polyfit(range(len(scores)), scores, 1)[0] if len(scores) >= 2 else 0
                slope = round(slope, 2)
                status = '不通过' if slope < thresholds[field] else '通过'
                if status == '不通过':
                    fail_count += 1
                ws.cell(row_idx, col_offset, slope)
                ws.cell(row_idx, col_offset + 1, status)
                col_offset += 2

            latest_year = max(results.keys()) if results else 0
            latest_score = results.get(latest_year, {}).get('总分', 0)
            if isinstance(latest_score, (int, float)) and latest_score < 55:
                final = '不通过'
            else:
                final = '不通过' if fail_count >= 2 else '通过'
            ws.cell(row_idx, col_offset, final)

            # ── 落库: 保存年度评分到 company_annual_scores ──
            # 本地取 ts_code(勿用闭包变量: 主循环里 ts_code 在 _write_results 之后才赋值,
            # 首个完成的股票会触发 "free variable referenced before assignment" 丢落库)
            _sc = str(ws.cell(row_idx, code_col).value or '').strip()
            try:
                _save_annual_scores_to_db(_sc, results, stock_years)
            except Exception as e:
                logger.debug(f'{_sc} 年度评分落库失败(不影响评分): {e}')

    # 行级续传：判定待补算行（三级行业为空 或 总分_T 缺失）；全量模式则所有行
    all_rows = list(range(2, ws.max_row + 1))
    l3_col = header_row.index('三级行业') + 1
    total_t_col = header_row.index('总分_T') + 1

    if resume_mode:
        def _needs_rescore(r):
            l3 = ws.cell(r, l3_col).value
            tt = ws.cell(r, total_t_col).value
            l3_empty = l3 is None or str(l3).strip() in ('', 'N/A', 'nan', 'None')
            tt_empty = tt is None or str(tt).strip() in ('', 'N/A', 'nan', 'None')
            return l3_empty or tt_empty
        row_indices = [r for r in all_rows if _needs_rescore(r)]
        logger.info(f"行级续传：已评分 {len(all_rows) - len(row_indices)} 行跳过，待补算 {len(row_indices)} 行")
    else:
        row_indices = all_rows
        logger.info(f"全量评分：共 {len(row_indices)} 行")

    if not row_indices:
        logger.info("✅ 无待补算行，直接保存退出（文件无变更）")
        wb.save(output_path)
        print(f"\n完成! 结果文件: {output_path} (无变更)")
        return

    _total = len(row_indices)

    # ── ③-d 加速: 批量预取比率(financial_indicators) + 行业(industry_data) + 持久阈值 ──
    # 替逐股 tushare 取数, 字节级等同原路径; 持久阈值跨重启复用。
    all_codes = []
    for r in row_indices:
        c = ws.cell(r, code_col).value
        if c:
            all_codes.append(str(c).strip())
    bulk_load_ratios(all_codes)
    bulk_load_industry()
    _load_persistent_thresholds()
    fi_cov = len({k[0] for k in _RATIOS_CACHE})
    logger.info(f"预取完成: 比率覆盖 {fi_cov}/{len(all_codes)} 股, 行业缓存 {len(_INDUSTRY_CACHE)} 股, "
                f"持久阈值 {len(_persistent_threshold_cache)} 条")

    MAX_WORKERS_SCORE = min(12, len(row_indices))
    logger.info(f"并发评分：{MAX_WORKERS_SCORE}线程并行")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS_SCORE) as executor:
        futures = {executor.submit(_score_one, ridx): ridx for ridx in row_indices}
        for future in as_completed(futures):
            ridx = futures[future]
            try:
                row_idx, results, stock_years = future.result()
                with _write_lock:
                    _write_results(row_idx, results, stock_years)
                    _done_count[0] += 1
                    n = _done_count[0]
                    if n % 10 == 0 or n == _total:
                        output_path = excel_path.replace('.xlsx', '_scored.xlsx')
                        try:
                            wb.save(output_path)
                        except Exception:
                            pass
                    ts_code = str(ws.cell(row_idx, code_col).value or '').strip()
                    logger.info(f"  [已评分 {n}/{_total}] {ts_code}")
            except Exception as e:
                logger.warning(f"  第{ridx}行评分失败: {e}")

    # 最终保存
    output_path = excel_path.replace('.xlsx', '_scored.xlsx')
    wb.save(output_path)
    logger.info(f"\n评分结果已保存: {output_path}")
    print(f"\n完成! 结果文件: {output_path}")




# ====== backfill_financial_indicators (indicators) ======
import argparse
import os
import sys
import time
import numpy as np
import pandas as pd
import pymysql
import tushare as ts

# 复用 EFAES 的财务比率算法(与定增训练集同源同公式)
EFAES_PATH = '/Users/davy/github/EFAES'
if EFAES_PATH not in sys.path:
    sys.path.insert(0, EFAES_PATH)
from src.core.calculate_financial_ratios import calculate_financial_ratios  # noqa: E402

# tushare token: 走 resolve_tushare_token(不再硬编码; 旧 literal f2380... 已泄漏, 需在 tushare.pro 轮换)
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from tushare_token import resolve_tushare_token  # noqa: E402
os.environ.setdefault('TUSHARE_TOKEN', resolve_tushare_token())

TABLE_FIELDS = [
    'current_ratio', 'quick_ratio', 'inv_turn', 'ar_turn', 'ca_turn', 'assets_turn',
    'roa', 'npta', 'roe', 'roe_dt', 'netprofit_margin', 'grossprofit_margin',
    'debt_to_assets', 'int_to_talcap', 'debt_to_eqt', 'ebit_to_interest',
    'cash_to_liqdebt', 'cash_to_liqdebt_withinterest', 'rd_exp_ratio',
    'op_yoy', 'ebt_yoy', 'netprofit_yoy', 'dt_netprofit_yoy', 'roe_yoy',
    'tr_yoy', 'or_yoy', 'equity_yoy',
]
# fina_indicator 不提供、必须从报表算的 5 个
NEED_STATEMENT = ['inv_turn', 'ebit_to_interest', 'cash_to_liqdebt',
                  'cash_to_liqdebt_withinterest', 'rd_exp_ratio']


def _fetch_statements(pro, code, start, end):
    """取三张表的年报行，附 report_year 列。"""
    out = {}
    for name, fn in [('balancesheet', pro.balancesheet),
                     ('income', pro.income),
                     ('cashflow', pro.cashflow)]:
        df = pd.DataFrame()
        for attempt in range(3):  # tushare 限频会返回空 → 退避重试
            try:
                df = fn(ts_code=code, start_date=start, end_date=end)
            except Exception:
                df = pd.DataFrame()
            if df is not None and len(df) > 0:
                break
            time.sleep(1.5 * (attempt + 1))
        time.sleep(0.25)
        if df is None or len(df) == 0 or 'end_date' not in df.columns:
            out[name] = pd.DataFrame(columns=['report_year'])
            continue
        ed = df['end_date'].astype(str)
        mask = ed.str.endswith('1231')
        df = df[mask].copy()
        df['report_year'] = ed[mask].str[:4].astype(int).values
        out[name] = df.reset_index(drop=True)
    return out


def main_indicators():
    ap = argparse.ArgumentParser(description='回填 financial_indicators(三表算法, 与定增同源)')
    ap.add_argument('excel', help='含 股票代码 列的 Excel')
    ap.add_argument('--token', default=None, help='tushare token(默认走 resolve_tushare_token)')
    ap.add_argument('--years', type=int, default=5)
    ap.add_argument('--start-year', type=int, default=2016)  # 多取几年供平均值(当年+上年)
    ap.add_argument('--end-year', type=int, default=2025)
    ap.add_argument('--force', action='store_true',
                    help='全量重算 Excel 每只(新股建行 + 既有行 DELETE+INSERT 覆盖); '
                         '不加则只补"既有行里 5 报表字段缺失"的股(回测全A新股原本无行, null-filter 会漏掉, 必加)')
    args = ap.parse_args()

    ex = pd.read_excel(args.excel, sheet_name='Sheet1')
    codes = [str(c) for c in ex['股票代码'].dropna().unique()]
    conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='',
                           database='investment_valuation', charset='utf8mb4')
    if args.force:
        # 全量: Excel 每只都重算(下方 INSERT 前先 DELETE, 幂等覆盖)。回测全A必走此路——
        # 新股原本无 financial_indicators 行, 默认 null-filter 只查"既有行缺字段"会整批漏掉。
        missing = codes
        print(f'--force 全量重算: Excel {len(codes)} 只(新股建行 + 既有行覆盖)')
    else:
        # 默认: 只补"既有行里 5 报表字段缺失"的(即只被 fina_indicator 快捷覆盖、缺 inv_turn/rd_exp_ratio)
        cs = ','.join([f"'{c}'" for c in codes])
        null_sql = (f"SELECT DISTINCT stock_code FROM financial_indicators "
                    f"WHERE stock_code IN ({cs}) AND (inv_turn IS NULL OR rd_exp_ratio IS NULL)")
        null_codes = set(pd.read_sql(null_sql, conn)['stock_code'])
        missing = [c for c in codes if c in null_codes]
        print(f'Excel {len(codes)} 只; 缺 5 报表字段待重算: {len(missing)}(回测全A新股请加 --force)')

    pro = ts.pro_api(args.token or os.environ['TUSHARE_TOKEN'])
    cur = conn.cursor()
    sd, ed = f'{args.start_year}0101', f'{args.end_year}1231'
    target_years = list(range(args.end_year - args.years + 1, args.end_year + 1))
    cols = ['stock_code', 'report_year'] + TABLE_FIELDS + ['used_average_values', 'valid_indicator_count', 'ann_date', 'end_date']
    ins_sql = f"INSERT INTO financial_indicators ({', '.join(cols)}) VALUES ({', '.join(['%s']*len(cols))})"

    ok = rows = fail = 0
    for i, code in enumerate(missing):
        try:
            stmts = _fetch_statements(pro, code, sd, ed)
            bs, isc, cf = stmts['balancesheet'], stmts['income'], stmts['cashflow']
            if len(bs) == 0 or len(isc) == 0:
                fail += 1
                continue
            computed = []  # (year, [27 vals], valid_count)
            for yr in target_years:
                try:
                    ratios = calculate_financial_ratios(bs, isc, cf, yr)
                except Exception:
                    continue
                if not ratios:
                    continue
                vals, vc = [], 0
                for f in TABLE_FIELDS:
                    v = ratios.get(f)
                    try:
                        v = float(v)
                    except (TypeError, ValueError):
                        v = None
                    if v is None or v != v or np.isinf(v):
                        v = None
                    else:
                        vc += 1
                    vals.append(v)
                if vc > 0:
                    is_row = isc[isc['report_year'] == yr]
                    ad = ed_ = None
                    if len(is_row) > 0:
                        ad = str(is_row.iloc[0].get('ann_date', ''))[:8]
                        ed_ = str(is_row.iloc[0].get('end_date', ''))[:8]
                        ad = ad if ad.isdigit() else None
                        ed_ = ed_ if ed_.isdigit() else None
                    computed.append((yr, vals, vc, ad, ed_))
            if not computed:
                fail += 1
                continue
            cur.execute('DELETE FROM financial_indicators WHERE stock_code=%s', (code,))
            for yr, vals, vc, ad, ed_ in computed:
                cur.execute(ins_sql, [code, yr] + vals + [0, vc, ad, ed_])
                rows += 1
            ok += 1
            if (i + 1) % 25 == 0:
                conn.commit()
                print(f'  进度 {i+1}/{len(missing)} ...')
        except Exception as e:
            fail += 1
            print(f'  {code} 失败: {e}')
    conn.commit()
    conn.close()
    print(f'完成: {ok}/{len(missing)} 只成功, 插入 {rows} 行, 失败 {fail}')




# ====== bulk_backfill_early_scores (early) ======
import os, sys, time, argparse, threading
sys.path.insert(0, os.path.expanduser('~/github/EFAES'))
from src.web.config import DB_CONFIG, TUSHARE_TOKEN
from src.core.calculate_financial_score import FinancialScoringCard
import tushare as ts, pandas as pd, numpy as np
from concurrent.futures import ThreadPoolExecutor

PRO = ts.pro_api(TUSHARE_TOKEN)
NEG = {'debt_to_assets', 'int_to_talcap', 'debt_to_eqt'}
INDICATORS = ['inv_turn', 'ar_turn', 'ca_turn', 'assets_turn', 'netprofit_margin', 'grossprofit_margin',
              'roe', 'roe_dt', 'roa', 'npta', 'current_ratio', 'quick_ratio', 'cash_to_liqdebt',
              'cash_to_liqdebt_withinterest', 'debt_to_assets', 'int_to_talcap', 'debt_to_eqt', 'ebit_to_interest',
              'netprofit_yoy', 'dt_netprofit_yoy', 'roe_yoy', 'tr_yoy', 'or_yoy', 'equity_yoy', 'op_yoy', 'ebt_yoy',
              'rd_exp_ratio']
_VIP_SEM = threading.Semaphore(4)  # vip 接口并发上限


def _vip(fn, period):
    for attempt in range(3):
        try:
            with _VIP_SEM:
                d = getattr(PRO, fn)(period=period)
            if d is not None and not d.empty:
                return d
        except Exception as e:
            print(f'    {fn}({period}) 重试{attempt}: {str(e)[:80]}')
            time.sleep(1.0 * (attempt + 1))
    return None


def bulk_indicators(years):
    """fina_indicator_vip 按期拉全市场财务指标(与 fina_indicator 同字段, vip 支持按期批量)。
    过滤年报(end_date==period), 按 ts_code 去重(保留最新 ann_date)。返回含 report_year 的 DataFrame。"""
    frames = []
    for y in years:
        d = _vip('fina_indicator_vip', f'{y}1231')
        if d is None or d.empty:
            print(f'  fina_indicator_vip({y}): 空')
            continue
        d = d[d['end_date'].astype(str) == f'{y}1231'].copy()
        if d.empty:
            continue
        d = d.sort_values('ann_date', ascending=False).drop_duplicates('ts_code', keep='first')
        d['report_year'] = y
        frames.append(d)
        print(f'  fina_indicator_vip({y}): {len(d)} 公司')
    ratios = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    if not ratios.empty:
        ratios['ts_code'] = ratios['ts_code'].astype(str)
        print(f'  合并: {len(ratios)} 行, {ratios["ts_code"].nunique()} 公司, '
              f'{ratios["report_year"].min()}~{ratios["report_year"].max()}年')
    return ratios


def industry_of(ts_code):
    """index_member_all 反查申万行业(支持并发, 退避重试)。返回名称+code(code 用于正向取成分)。"""
    for attempt in range(3):
        try:
            d = PRO.index_member_all(ts_code=ts_code)
            if d is not None and not d.empty:
                r = d.sort_values('in_date', ascending=False).iloc[0]
                if r.get('l3_name'):
                    return {'l3': r.get('l3_name'), 'l3_code': r.get('l3_code'),
                            'l2': r.get('l2_name'), 'l2_code': r.get('l2_code'),
                            'l1': r.get('l1_name'), 'l1_code': r.get('l1_code')}
                return None  # 非空但无三级 → 真无分类, 不重试
        except Exception:
            pass
        if attempt < 2:
            time.sleep(0.5 * (attempt + 1))
    return None


def members_by_code(code, level):
    """index_member_all 正向(按 l3_code/l2_code)取行业成分股 ts_code, 带重试。
    index_member_all 单接口双向: 反查用 ts_code, 正查用 l3_code/l2_code(注意用 code 不是 name,
    name 不被识别会返回全量)。并发友好, 替代 index_classify+index_member 两步。"""
    if not code:
        return []
    for attempt in range(3):
        try:
            d = PRO.index_member_all(**{f'{level}_code': code})
            if d is not None and not d.empty:
                return d['ts_code'].astype(str).unique().tolist()
            return []  # 非空但无成分 → 真无
        except Exception:
            if attempt < 2:
                time.sleep(0.6 * (attempt + 1))
    return []


def compute_thresholds(ratios, industry_members_map):
    """按 (行业, 年) 出阈值 {indicator: {极差/较差/中等/良好/优秀}}。负向指标反向。"""
    thresholds = {}
    for l3, members in industry_members_map.items():
        if not members:
            continue
        sub = ratios[ratios['ts_code'].isin(members)]
        for y, g in sub.groupby('report_year'):
            if len(g) < 5:
                continue
            t = {}
            for col in INDICATORS:
                if col not in g.columns:
                    continue
                vals = g[col].replace([np.inf, -np.inf], np.nan).dropna()
                if len(vals) < 5:
                    continue
                p10, p30, p50, p70, p90 = (vals.quantile(q) for q in (0.1, 0.3, 0.5, 0.7, 0.9))
                if pd.isna(p50):
                    continue
                if col in NEG:
                    t[col] = {'优秀': p10, '良好': p30, '中等': p50, '较差': p70, '极差': p90}
                else:
                    t[col] = {'极差': p10, '较差': p30, '中等': p50, '良好': p70, '优秀': p90}
            if len(t) >= 10:
                thresholds[(l3, int(y))] = t
    return thresholds


def score_stock(ts_code, t_year, ind_info, ratios, thr_l3, thr_l2, scorer):
    """本地评一只股 T-4..T 年。阈值三级→二级 回退(冷门三级样本不足时兜底)。"""
    results = {}
    l3 = ind_info.get('l3') if ind_info else None
    l2 = ind_info.get('l2') if ind_info else None
    for y in range(t_year - 4, t_year + 1):
        rrow = ratios[(ratios['ts_code'] == ts_code) & (ratios['report_year'] == y)]
        thr = thr_l3.get((l3, y)) or thr_l2.get((l2, y))
        if rrow.empty or not thr:
            continue
        ratios_dict = {k: v for k, v in rrow.iloc[0].to_dict().items()
                       if k not in ('ts_code', 'company_name', 'ann_date', 'end_date', 'report_year', 'used_average_values')}
        scorer.set_industry_thresholds(thr)
        scorer.load_company_data(ratios_dict, ts_code)
        sc = scorer.calculate_score()
        results[y] = {'总分': round(sc['总得分'], 2), '评级': scorer.get_rating(sc['总得分']),
                      '盈利能力': round(sc['维度得分'].get('盈利能力', 0), 2),
                      '成长能力': round(sc['维度得分'].get('成长能力', 0), 2),
                      '运营能力': round(sc['维度得分'].get('运营能力', 0), 2),
                      '偿债能力': round(sc['维度得分'].get('偿债能力', 0), 2)}
    return results


def save_to_db(stock_code, ind_info, results):
    import pymysql
    if not results:
        return
    conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='',
                           database='investment_valuation', charset='utf8mb4')
    try:
        with conn.cursor() as cur:
            for year, r in results.items():
                cur.execute("""INSERT INTO company_annual_scores
                    (stock_code, report_year, total_score, rating, profitability, growth, operating, solvency,
                     industry_l1, industry_l2, industry_l3)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                    ON DUPLICATE KEY UPDATE total_score=VALUES(total_score), rating=VALUES(rating),
                    profitability=VALUES(profitability), growth=VALUES(growth), operating=VALUES(operating),
                    solvency=VALUES(solvency), industry_l1=VALUES(industry_l1), industry_l2=VALUES(industry_l2),
                    industry_l3=VALUES(industry_l3)""",
                            (stock_code, year, r['总分'], r['评级'], r['盈利能力'], r['成长能力'], r['运营能力'], r['偿债能力'],
                             (ind_info or {}).get('l1'), (ind_info or {}).get('l2'), (ind_info or {}).get('l3')))
            conn.commit()
    finally:
        conn.close()


def main_early():
    ap = argparse.ArgumentParser(description='Bulk vip 回填 2010-2016 年度评分')
    ap.add_argument('excel', help='输入 Excel(股票代码/股票简称/报价日)')
    ap.add_argument('--limit', type=int, default=0, help='只处理前N只(0=全部)')
    args = ap.parse_args()

    df = pd.read_excel(args.excel)
    df = df.rename(columns={c: c.strip() for c in df.columns})
    placements = [(str(r['股票代码']).strip(), int(str(r['报价日']).strip()[:4]))
                  for _, r in df.iterrows() if pd.notna(r['股票代码'])]
    if args.limit:
        placements = placements[:args.limit]
    print(f'待回填: {len(placements)} 只股')

    # 年份范围(报价日 T-4..T 的并集; +1年做增长率基期)
    yrs = sorted({t - 4 for _, t in placements} | {t for _, t in placements})
    fetch_years = list(range(min(yrs) - 1, max(yrs) + 1))
    print(f'取数年份(含增长基期): {fetch_years[0]}~{fetch_years[-1]}')

    print('\n[1/3] fina_indicator_vip 按期拉全市场指标...')
    ratios = bulk_indicators(fetch_years)
    ratios['report_year'] = ratios['report_year'].astype(int)

    print('\n[2/3] 各股行业反查 + 三级/二级成分股 + 全部(行业,年)阈值...')
    unique_codes = sorted({c for c, _ in placements})
    with ThreadPoolExecutor(max_workers=8) as ex:
        ind_map = dict(ex.map(lambda c: (c, industry_of(c)), unique_codes))
    print(f'  {len(unique_codes)} 只唯一股 / {len(placements)} 条placement')
    unique_l3 = sorted({i['l3'] for i in ind_map.values() if i and i.get('l3')})
    unique_l2 = sorted({i['l2'] for i in ind_map.values() if i and i.get('l2')})
    print(f'  涉及 {len(unique_l3)} 个三级 / {len(unique_l2)} 个二级行业')

    def _code_for(name, name_key, code_key):
        return next((i.get(code_key) for i in ind_map.values() if i and i.get(name_key) == name), None)

    with ThreadPoolExecutor(max_workers=8) as ex:
        mm_l3 = dict(zip(unique_l3, ex.map(lambda n: members_by_code(_code_for(n, 'l3', 'l3_code'), 'l3'), unique_l3)))
        mm_l2 = dict(zip(unique_l2, ex.map(lambda n: members_by_code(_code_for(n, 'l2', 'l2_code'), 'l2'), unique_l2)))
    thr_l3 = compute_thresholds(ratios, mm_l3)
    thr_l2 = compute_thresholds(ratios, mm_l2)
    print(f'  阈值: 三级 {len(thr_l3)} / 二级 {len(thr_l2)} 个(行业,年)组合')

    print('\n[3/3] 本地评分 + 落库...')
    scorer = FinancialScoringCard(DB_CONFIG)
    ok = 0
    for i, (code, t_year) in enumerate(placements, 1):
        ind = ind_map.get(code)
        if not ind or not (ind.get('l3') or ind.get('l2')):
            continue
        results = score_stock(code, t_year, ind, ratios, thr_l3, thr_l2, scorer)
        if results:
            save_to_db(code, ind, results)
            ok += 1
        if i % 50 == 0 or i == len(placements):
            print(f'  [{i}/{len(placements)}] 已评分 {ok} 只')
    print(f'\n完成: {ok}/{len(placements)} 只股评分落库')




# ====== backfill_ann_date (ann_date) ======
import argparse
import os
import sys
import time
import pymysql
import tushare as ts

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from tushare_token import resolve_tushare_token   # noqa: E402
os.environ.setdefault('TUSHARE_TOKEN', resolve_tushare_token())


def main_ann_date():
    ap = argparse.ArgumentParser()
    ap.add_argument('--token', default=None, help='tushare token(默认走 resolve_tushare_token)')
    ap.add_argument('--sleep', type=float, default=0.18)
    args = ap.parse_args()

    conn = pymysql.connect(host='127.0.0.1', port=3306, user='root', password='',
                           database='investment_valuation', charset='utf8mb4')
    cur = conn.cursor()
    cur.execute('SELECT DISTINCT stock_code FROM financial_indicators WHERE ann_date IS NULL')
    codes = [r[0] for r in cur.fetchall()]
    print(f'待回填 ann_date 的股票: {len(codes)}')

    pro = ts.pro_api(args.token or os.environ['TUSHARE_TOKEN'])
    ok = rows = fail = 0
    for i, code in enumerate(codes):
        try:
            df = None
            for attempt in range(3):  # 限频退避
                try:
                    df = pro.fina_indicator(ts_code=code, start_date='20140101', end_date='20251231')
                except Exception:
                    df = None
                if df is not None and len(df) > 0:
                    break
                time.sleep(1.2 * (attempt + 1))
            time.sleep(args.sleep)
            if df is None or len(df) == 0 or 'ann_date' not in df.columns:
                fail += 1
                continue
            # 取年报(end_date 以 1231 结尾), 建 report_year -> (ann_date, end_date)
            ann_map = {}
            for _, r in df.iterrows():
                ed = str(r.get('end_date', ''))
                if not ed.endswith('1231'):
                    continue
                yr = int(ed[:4])
                ad = str(r.get('ann_date', '') or '')
                if ad and ad != 'nan':
                    ann_map[yr] = (ad[:8], ed[:8])
            if not ann_map:
                fail += 1
                continue
            # UPDATE 该股票对应 report_year 的 ann_date/end_date
            for yr, (ad, ed) in ann_map.items():
                cur.execute(
                    'UPDATE financial_indicators SET ann_date=%s, end_date=%s WHERE stock_code=%s AND report_year=%s',
                    (ad, ed, code, yr))
                rows += cur.rowcount
            conn.commit()
            ok += 1
            if (i + 1) % 50 == 0:
                print(f'  进度 {i+1}/{len(codes)} ...')
        except Exception as e:
            fail += 1
            print(f'  {code} 失败: {e}')
    conn.close()
    print(f'完成: {ok}/{len(codes)} 只成功, 更新 {rows} 行, 失败 {fail}')



if __name__ == '__main__':
    _cmd=sys.argv[1] if len(sys.argv)>1 else 'score'
    sys.argv=[sys.argv[0]]+sys.argv[2:]
    {'score': main_score, 'indicators': main_indicators, 'early': main_early, 'ann_date': main_ann_date}[_cmd]()
