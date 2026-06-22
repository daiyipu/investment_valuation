#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
定增盈利概率预测 - 加载训练好的模型对新样本预测

模型版本由 model_registry 的 current.full / current.scorecard 指针决定
（见 manage_models.py current）。不再写死具体版本/AUC。

支持两种用法:
  1. 独立运行: python ml_training/predict_profitability.py <scored_excel> [--output result.xlsx]
  2. 被调用:   from predict_profitability import predict; df = predict(scored_excel_path)

输出两列:
  - 盈利概率_LightGBM   (current.full 模型概率; SC 模型时即评分卡概率)
  - 盈利概率_逻辑回归   (LGB 模型: LR 概率; SC 模型: 同评分卡概率)
"""

import sys
import os
import re
import json
import warnings
import pickle
import numpy as np
import pandas as pd

warnings.filterwarnings('ignore')

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, os.path.join(SCRIPT_DIR, 'pipeline'))   # 管线模块已移入 pipeline/


_INTERVAL_RE = re.compile(r'^[\[(](.*?),\s*(.*?)[\])]$')


def _parse_interval(key):
    """解析 pandas 区间字符串 '(0.33, 2.145]' → (0.33, 2.145)。失败返回 None。"""
    m = _INTERVAL_RE.match(str(key).strip())
    if not m:
        return None
    try:
        return float(m.group(1)), float(m.group(2))
    except ValueError:
        return None


def score_with_scorecard(df, sc_dir):
    """用评分卡模型对 df 每行打分。

    score = base_points + Σ(B · coef_i · woe_i)
    缺失/越界特征按最近 bin 兜底；NaN 贡献 0（与训练时 evaluate_scorecard 的
    fillna(0) 行为一致）。

    Args:
        df: 已构造完整特征的 DataFrame
        sc_dir: 评分卡版本目录（含 scorecard_model.pkl）

    Returns:
        np.array[float] 每行得分
    """
    with open(os.path.join(sc_dir, 'scorecard_model.pkl'), 'rb') as f:
        sc = pickle.load(f)
    model = sc['model']
    features = sc['features']
    woe_bins = sc['woe_bins']
    base_points = float(sc['scoring_params']['base_points'])
    B = float(sc['scoring_params']['B'])
    coefs = dict(zip(features, model.coef_[0]))

    scores = np.full(len(df), base_points, dtype=float)
    for feat in features:
        if feat not in df.columns or feat not in woe_bins:
            continue
        # 预解析该特征的 bins: [(left, right, woe), ...] 按 left 排序
        parsed = []
        for k, woe in woe_bins[feat].get('woe_map', {}).items():
            iv = _parse_interval(k)
            if iv is not None:
                parsed.append((iv[0], iv[1], float(woe)))
        if not parsed:
            continue
        parsed.sort(key=lambda x: x[0])
        min_left = parsed[0][0]
        first_woe = parsed[0][2]
        last_woe = parsed[-1][2]

        vals = pd.to_numeric(df[feat], errors='coerce').values
        coef = coefs[feat]
        for i, v in enumerate(vals):
            if v != v:  # NaN → 跳过（贡献 0）
                continue
            woe = None
            for l, r, w in parsed:
                if l < v <= r:
                    woe = w
                    break
            if woe is None:
                woe = first_woe if v <= min_left else last_woe  # 越界兜底
            scores[i] += B * coef * woe
    return scores


def predict(scored_excel_path):
    """
    对 scored Excel 中的每只股票预测盈利概率

    流程:
      1. 从 Excel 提取财务评分/子场景等特征
      2. 从 MySQL 加载行情/估值/FCF 等DB特征
      3. 计算衍生特征（FCF增长率、交叉比率、评分变动、估值相对、行情动量、行业/大盘）
      4. 加载全量特征模型的 meta（特征列表 + median）
      5. LightGBM + 逻辑回归 双模型预测

    Args:
        scored_excel_path: 已评分的 Excel 文件路径（batch_screen_and_score 输出）

    Returns:
        DataFrame: SC 模型→[股票代码, 盈利概率_评分卡]; LGB 模型→[股票代码, 盈利概率_LightGBM, 盈利概率_逻辑回归]
    """
    import lightgbm as lgb
    from export_features import load_db_features, load_scored_features, load_financial_ratios
    from derive_features import (
        derive_fcf_growth_rates, derive_fcf_cross_metrics,
        derive_financial_score_deltas, derive_valuation_relative,
        derive_market_momentum, derive_industry_valuation_growth,
        derive_market_index_features,
    )
    from model_registry import require_current_dir, get_current
    from db_model_store import load_predict_bundle   # full 模型权重+meta 从 DB 加载(不再读磁盘 version 目录)

    output_dir = os.path.join(SCRIPT_DIR, 'output')
    # 模型从 registry 的当前生产版本读取（见 manage_models.py current）
    version = get_current('full')
    bundle = load_predict_bundle(version)   # 权重+features+medians 从 DB ml_model_meta 加载
    print(f'  [ML-0] 使用 full 模型版本: {version} (权重从 DB 加载)')
    # 旧的独立评分卡得分列([ML-8])已移除: current.full 现在本身就是评分卡(SC)模型,
    # 主概率列即为评分卡概率, 不再另设老的 on-disk scorecard 得分列。

    # ═══════════════════════════════════════════════
    # 1. 从 Excel 提取特征
    # ═══════════════════════════════════════════════
    print('\n  [ML-1] 加载 Excel 特征...')
    scored = load_scored_features(scored_excel_path)
    stock_codes = scored['股票代码'].tolist()
    print(f'    样本: {len(scored)} 条')

    # ═══════════════════════════════════════════════
    # 2. 从 DB 加载特征
    # ═══════════════════════════════════════════════
    print('  [ML-2] 加载 DB 特征...')
    sample_keys = []
    for _, row in scored.iterrows():
        code = row['股票代码']
        issue_date = str(row.get('报价日', '')).replace('.0', '').strip()
        if not issue_date or issue_date == 'nan' or len(issue_date) < 8:
            issue_date = None
        sample_keys.append((code, issue_date))

    db_feats = load_db_features(sample_keys)
    matched_price = db_feats.get('当前价', pd.Series()).notna().sum()
    print(f'    行情匹配: {matched_price}/{len(scored)}')

    # 财务比率（financial_indicators API；三表已弃用，覆盖率0.43%且与比率表重复）
    try:
        ratio_feats = load_financial_ratios(sample_keys)
    except Exception as e:
        print(f'    财务比率跳过: {e}')
        ratio_feats = pd.DataFrame()

    # ═══════════════════════════════════════════════
    # 3. 合并
    # ═══════════════════════════════════════════════
    scored = scored.reset_index(drop=True)
    # 去掉 db_feats 中与 scored 同名的列（报价日/定增决策/报价日价格…），
    # 一律取 scored（权威源），避免 concat 后出现重名列导致 df[c] 变成多列。
    _dup_cols = (set(scored.columns) & set(db_feats.columns)) - {'股票代码'}
    if _dup_cols:
        print(f'    去重列(取scored版): {sorted(_dup_cols)}')
    db_feat_cols = [c for c in db_feats.columns if c != '股票代码' and c not in _dup_cols]
    df = pd.concat([scored, db_feats[db_feat_cols].reset_index(drop=True)], axis=1)

    if not ratio_feats.empty:
        df = pd.concat([df.reset_index(drop=True), ratio_feats.reset_index(drop=True)], axis=1)

    # 清理类型
    str_keep = {'股票代码', '股票简称', '最终结论', '一级行业', '二级行业', '三级行业',
                '定价方式', '定增决策', '行业代码', '行业名称'}
    for c in df.columns:
        if c in str_keep:
            df[c] = df[c].astype(str)
        elif df[c].dtype == object or str(df[c].dtype) == 'category':
            df[c] = pd.to_numeric(df[c], errors='coerce')

    # ═══════════════════════════════════════════════
    # 4. 衍生特征
    # ═══════════════════════════════════════════════
    print('  [ML-3] 衍生特征...')
    for func in [derive_fcf_growth_rates, derive_fcf_cross_metrics,
                 derive_financial_score_deltas, derive_valuation_relative,
                 derive_market_momentum]:
        df = func(df)
    try:
        df = derive_industry_valuation_growth(df)
        df = derive_market_index_features(df)
    except Exception as e:
        print(f'    行业/大盘衍生跳过: {e}')

    # ── 因子引擎 + 策略信号(与 derive_features 完全一致, 保证训练/预测特征空间统一) ──
    try:
        from derive_features import derive_strategy_signals, derive_alpha_beta_factors
        df = derive_strategy_signals(df)
        df = derive_alpha_beta_factors(df)
        print(f'    因子引擎: +{sum(1 for c in df.columns if c.startswith(("MACD","RSI","KDJ","BOLL","ROC_","k_","ret_","vol_","amount_","corr_","vwap_","beta_","idiovol","三浪","抵抗")))} 因子列')
    except Exception as e:
        import traceback; traceback.print_exc()
        print(f'    ⚠ 因子引擎跳过: {e}')

    df = df.replace([np.inf, -np.inf], np.nan)
    print(f'    最终特征: {df.shape[1]} 列')

    # ═══════════════════════════════════════════════
    # 5. 加载全量特征模型的 meta
    # ═══════════════════════════════════════════════
    print('  [ML-4] 加载模型 meta(从 DB)...')
    model_features = bundle['features']
    train_medians = bundle['medians']
    print(f'    模型特征: {len(model_features)} 个')

    # ═══════════════════════════════════════════════
    # 6. 构建预测特征矩阵
    # ═══════════════════════════════════════════════
    print('  [ML-5] 构建预测特征矩阵...')
    X_pred = pd.DataFrame(index=df.index)
    missing_list = []
    for feat in model_features:
        if feat in df.columns:
            X_pred[feat] = pd.to_numeric(df[feat], errors='coerce')
        else:
            X_pred[feat] = np.nan
            missing_list.append(feat)

    # 用训练集 median 填充
    for feat in model_features:
        X_pred[feat] = X_pred[feat].fillna(train_medians.get(feat, 0))
    X_pred = X_pred.replace([np.inf, -np.inf], 0).fillna(0)

    if missing_list:
        print(f'    ⚠️ 缺少 {len(missing_list)} 个特征(median填充): {missing_list[:5]}...')

    # 特征覆盖率
    total_cells = len(X_pred) * len(model_features)
    non_null = sum(df[f].notna().sum() for f in model_features if f in df.columns)
    print(f'    特征覆盖率(填充前): {non_null/total_cells*100:.1f}%')

    # ═══════════════════════════════════════════════
    # 7/8. 预测(LGB+LR 或 评分卡 SC 二选一)
    # ═══════════════════════════════════════════════
    is_sc = not bundle.get('lgb_model')   # 评分卡模型: lgb_model 为空
    if is_sc:
        # 评分卡路径: WOE 分箱 + LR 打分(替代 LGB+LR)
        print('  [ML-6/7] 评分卡(SC) 预测(WOE+LR)...')
        from eval_loyo import apply_woe
        sc = pickle.loads(bundle['lr_bundle'])   # {kind, woe_bins, lr_model, features, medians}
        X_woe = apply_woe(X_pred, sc['features'], sc['woe_bins']).replace([np.inf, -np.inf], 0).fillna(0)
        lgb_proba = sc['lr_model'].predict_proba(X_woe)[:, 1]
        lr_proba = lgb_proba                      # SC 即主分数, 无独立 LR
        print(f'    SC 概率范围: [{lgb_proba.min():.3f}, {lgb_proba.max():.3f}]')
    else:
        print('  [ML-6] LightGBM 预测...')
        booster = lgb.Booster(model_str=bundle['lgb_model'])
        lgb_feature_names = booster.feature_name()
        X_lgb = pd.DataFrame(index=X_pred.index)
        for feat in lgb_feature_names:
            X_lgb[feat] = X_pred[feat] if feat in X_pred.columns else 0
        X_lgb = X_lgb.astype(float).values
        lgb_proba = booster.predict(X_lgb)
        print(f'    概率范围: [{lgb_proba.min():.3f}, {lgb_proba.max():.3f}]')

        print('  [ML-7] 逻辑回归 预测...')
        lr_data = pickle.loads(bundle['lr_bundle'])
        lr_model, lr_scaler, lr_feature_names = lr_data['model'], lr_data['scaler'], lr_data['features']
        X_lr = pd.DataFrame(index=X_pred.index)
        for feat in lr_feature_names:
            X_lr[feat] = X_pred[feat] if feat in X_pred.columns else 0
        X_lr = X_lr.replace([np.inf, -np.inf], 0).fillna(0)
        lr_proba = lr_model.predict_proba(lr_scaler.transform(X_lr.values))[:, 1]
        print(f'    概率范围: [{lr_proba.min():.3f}, {lr_proba.max():.3f}]')

    # ═══════════════════════════════════════════════
    # 9. 返回(按模型类型出列: SC→单列评分卡概率; LGB→LightGBM+逻辑回归两列)
    #    每个概率列配一个 1-10 档位(10=该批最高概率档, 便于挑高概率): rank(pct)*10 向上取整
    # ═══════════════════════════════════════════════
    result = pd.DataFrame({'股票代码': df['股票代码']})

    def _score_sc(sc_bundle, sc_features, sc_medians, sc_deciles, label_tag, X_src_df):
        """用一套 SC 模型打分, 返回 {盈利概率_tag, 档位_tag} 两列 + 概率数组。"""
        from eval_loyo import apply_woe as _aw
        X_sc = pd.DataFrame(index=X_src_df.index)
        for feat in sc_features:
            X_sc[feat] = pd.to_numeric(X_src_df[feat], errors='coerce') if feat in X_src_df.columns else np.nan
        X_sc = X_sc.fillna(sc_medians).replace([np.inf, -np.inf], 0)
        Xw = _aw(X_sc, sc_features, sc_bundle['woe_bins']).replace([np.inf, -np.inf], 0).fillna(0)
        proba = sc_bundle['lr_model'].predict_proba(Xw)[:, 1]
        if sc_deciles:
            tier = np.clip(np.searchsorted(sc_deciles, proba, side='right') + 1, 1, 10)
        else:
            tier = np.clip(np.ceil(pd.Series(proba).rank(pct=True) * 10), 1, 10)
        return proba, tier.astype(int)

    # 主模型(当前生产)
    if is_sc:
        sc_main = pickle.loads(bundle['lr_bundle'])
        main_feats = sc_main['features']
        main_deciles = sc_main.get('proba_deciles')
        main_tag = f'评分卡({len(main_feats)}特征)'
        p_main, t_main = _score_sc(sc_main, main_feats, sc_main.get('medians', {}),
                                    main_deciles, main_tag, df)
        result[f'盈利概率_{main_tag}'] = p_main
        result[f'档位_{main_tag}'] = t_main
    else:
        main_tag = 'LightGBM'
        result[f'盈利概率_{main_tag}'] = lgb_proba
        result[f'档位_{main_tag}'] = np.clip(np.ceil(pd.Series(lgb_proba).rank(pct=True) * 10), 1, 10).astype(int)

    # ═══════════════════════════════════════════════
    # 9b. 蓝绿对比: 同时用上一个生产版本(BLUE)打分
    # ═══════════════════════════════════════════════
    try:
        from db_model_store import list_model_metas, get_model_meta
        from model_registry import get_previous
        all_metas = list_model_metas(kind='gray')
        cur_nfeat = len(sc_main['features']) if is_sc else 0
        _cur_h = (get_model_meta(version) or {}).get('horizon')   # current 期限(如 7m)
        # 蓝绿 BLUE 优先 = 上一生产版(previous=老生产版本, 做对比+回滚); 不匹配则回退"最不同体量同期限SC"
        blue_ver = get_previous('full')
        bm = get_model_meta(blue_ver) if blue_ver else None
        blue_ok = (bm and blue_ver != version and not bm.get('lgb_model')
                   and bm.get('kind') == 'gray'
                   and (_cur_h is None or bm.get('horizon') == _cur_h))
        if not blue_ok:
            blue_candidates = [m for m in all_metas
                               if m['version'] != version and not m.get('lgb_model')
                               and m.get('kind') == 'gray'
                               and (_cur_h is None or m.get('horizon') == _cur_h)]
            if blue_candidates:
                blue_candidates.sort(key=lambda m: abs(m.get('n_features', 0) - cur_nfeat), reverse=True)
                blue_ver = blue_candidates[0]['version']
            else:
                blue_ver = None
        if blue_ver:
            blue_bundle_data = load_predict_bundle(blue_ver)
            sc_blue = pickle.loads(blue_bundle_data['lr_bundle'])
            blue_feats = sc_blue['features']
            blue_deciles = sc_blue.get('proba_deciles')
            blue_tag = f'评分卡BLUE({len(blue_feats)}特征)'
            p_blue, t_blue = _score_sc(sc_blue, blue_feats, sc_blue.get('medians', {}),
                                        blue_deciles, blue_tag, df)
            result[f'盈利概率_{blue_tag}'] = p_blue
            result[f'档位_{blue_tag}'] = t_blue
            print(f'  🔄 蓝绿对比: 同时输出 BLUE({blue_ver}, {len(blue_feats)}特征)')
    except Exception as e:
        print(f'  (无蓝绿对比: {e})')

    # ═══════════════════════════════════════════════
    # 9c. 多期限短期补充列: 1m / 3m / 1w / 2w 灰度 SC(7m 生产之外的短期价格维持概率)
    #     各期限生产 gray 标签(GRAY_CFG sweep 定: 1m·3m±10 / 1w±5 / 2w±6), 同套 _score_sc 打分;
    #     模型由 train_to_production.py --horizon 1/3/1w/2w --kind gray 训。DB 找不到最新版则跳过。
    # ═══════════════════════════════════════════════
    try:
        from db_model_store import list_model_metas as _lmm
    except Exception:
        _lmm = None
    if _lmm is not None:
        for _horizon, _lc, _tag in ((1, '1m_gray_sc', '1m'), (3, '3m_gray_sc', '3m'),
                                    ('1w', '1w_gray_sc', '1w'), ('2w', '2w_gray_sc', '2w')):
            try:
                _cands = [m for m in _lmm(label_config=_lc) if not m.get('lgb_model')]
                if not _cands:
                    continue
                _m = _cands[0]                       # created_at DESC → 最新版
                _sc = pickle.loads(_m['lr_bundle'])
                _feats = _sc['features']
                _p, _t = _score_sc(_sc, _feats, _sc.get('medians', {}),
                                   _sc.get('proba_deciles'), _tag, df)
                result[f'盈利概率_{_tag}'] = _p
                result[f'档位_{_tag}'] = _t
                _mt = _m.get('metrics', {})
                _loyo = _mt.get('sc_loyo_auc')
                if isinstance(_loyo, float) and _loyo == _loyo:
                    _perf = f', LOYO AUC={_loyo:.3f}±{_mt.get("sc_loyo_auc_std", 0):.3f}'
                else:
                    _fit = _mt.get('sc_fit_auc') or _mt.get('sc_oot_auc')
                    _perf = f', 自评AUC={_fit:.3f}(含泄漏)' if isinstance(_fit, float) else ''
                print(f'  ➕ 短期补充 {_tag}灰度SC: {_m["version"]} ({len(_feats)}特征{_perf})')
            except Exception as e:
                print(f'  ({_tag}短期列跳过: {e})')

    print(f'  ✅ ML预测完成: {len(result)} 条\n')
    return result


def write_to_excel(scored_excel_path, result_df, output_path=None):
    """将预测结果写入 Excel（追加到最后两列，按行序一一对应写入）

    注意: 同一股票可能多次定增，因此不能按股票代码映射，
    必须按行序逐行写入（result_df 与 Excel 行数一致、顺序一致）。
    """
    import openpyxl

    target = output_path or scored_excel_path
    wb = openpyxl.load_workbook(target)
    ws = wb.active

    # 待写入列(除股票代码); 幂等: 先删所有 盈利概率_* / 档位_* 列(防多次运行残留, 含 1m/3m 等任意期限)
    out_cols = [c for c in result_df.columns if c != '股票代码']
    while True:
        header_row = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
        tgt = next((h for h in header_row if isinstance(h, str)
                    and (h.startswith('盈利概率_') or h.startswith('档位_'))), None)
        if tgt is None:
            break
        ws.delete_cols(header_row.index(tgt) + 1)

    # 依次写入列
    col_map = {}
    for col_name in out_cols:
        c = ws.max_column + 1
        ws.cell(1, c, col_name)
        col_map[col_name] = c

    # 按行序写入（row 2 = 第1条数据, 与 result_df 第0行对应）: 概率列写%, 档位列写整数
    matched = 0
    data_rows = ws.max_row - 1  # 排除 header
    n = min(len(result_df), data_rows)
    for i in range(n):
        r = i + 2  # Excel 行号
        for col_name in out_cols:
            val = result_df.iloc[i][col_name]
            cell = f'{val*100:.1f}%' if col_name.startswith('盈利概率') else int(val)
            ws.cell(r, col_map[col_name], cell)
        matched += 1

    wb.save(target)
    print(f'  ✅ 写入 Excel: {matched} 条 → {target} (列: {", ".join(out_cols)})')
    return matched


def main():
    import argparse
    parser = argparse.ArgumentParser(description='定增盈利概率预测（双模型·全量特征）')
    parser.add_argument('scored_excel', help='已评分 Excel 路径')
    parser.add_argument('--output', default=None, help='输出路径（默认覆盖原文件）')
    args = parser.parse_args()

    print('=' * 60)
    print('定增盈利概率预测(按 current.full 模型类型: SC→评分卡 / LGB→LightGBM+逻辑回归)')
    print('=' * 60)

    result_df = predict(args.scored_excel)

    # 汇总统计(概率列: 均值/>50%; 档位列: 高分档分布, 便于挑高概率)
    print('\n  汇总:')
    for col in [c for c in result_df.columns if c != '股票代码']:
        if col.startswith('盈利概率'):
            m = result_df[col].mean()
            hi = (result_df[col] > 0.5).sum()
            print(f'    {col}: 均值 {m*100:.1f}% | >50% {hi} ({hi/len(result_df)*100:.1f}%)')
        else:   # 档位列(1-10, 10最高)
            dist = result_df[col].value_counts().sort_index(ascending=False)
            top = '  '.join(f'{t}档:{int(dist.get(t, 0))}' for t in range(10, 6, -1))
            high = int(result_df[col].ge(8).sum())
            print(f'    {col}: {top}  (8-10档共 {high} 个, 占 {high/len(result_df)*100:.1f}%)')

    # 写入 Excel
    write_to_excel(args.scored_excel, result_df, args.output)

    print('\n' + '=' * 60)
    print('🎉 预测完成!')
    print('=' * 60)


if __name__ == '__main__':
    main()
